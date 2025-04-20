import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_construction_bidding_workbook(filename="Construction_Bidding_Workbook.xlsx"):
    """
    Creates a comprehensive Excel workbook for construction bidding with:
    - Prequalification Form
    - Project Bid Package
    - Bid Form
    - Instructions to Bidders
    - Bid Manual TOC
    """
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    section_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== PREQUALIFICATION FORM =====
    prequalification = wb.create_sheet("Prequalification Form")
    
    # Set column widths
    prequalification.column_dimensions['A'].width = 30
    prequalification.column_dimensions['B'].width = 20
    prequalification.column_dimensions['C'].width = 20
    prequalification.column_dimensions['D'].width = 20
    prequalification.column_dimensions['E'].width = 20
    
    # Title
    prequalification['A1'] = "CONTRACTOR PREQUALIFICATION FORM"
    prequalification['A1'].font = Font(name='Arial', size=14, bold=True)
    prequalification.merge_cells('A1:E1')
    prequalification['A1'].alignment = Alignment(horizontal='center')
    
    # Add sections
    sections = [
        (3, "COMPANY INFORMATION"),
        (13, "CONTACT INFORMATION"),
        (23, "COMPANY DETAILS"),
        (31, "LICENSES & CERTIFICATIONS"),
        (40, "INSURANCE INFORMATION"),
        (50, "EXPERIENCE"),
        (57, "REFERENCES"),
        (78, "SAFETY"),
        (83, "LEGAL"),
        (89, "SIGNATURE")
    ]
    
    for row, title in sections:
        prequalification[f'A{row}'] = title
        prequalification[f'A{row}'].font = header_font
        prequalification[f'A{row}'].fill = header_fill
        prequalification.merge_cells(f'A{row}:E{row}')
    
    # Company Information
    company_fields = [
        "Company Name:", "Address:", "City, State, ZIP:", "Phone Number:", 
        "Email:", "Website:", "Year Established:", "Federal Tax ID:"
    ]
    
    for i, field in enumerate(company_fields, 4):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
    
    # Contact Information
    contact_fields = [
        "Primary Contact:", "Title:", "Phone:", "Email:",
        "Secondary Contact:", "Title:", "Phone:", "Email:"
    ]
    
    for i, field in enumerate(contact_fields, 14):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
    
    # Company Details
    prequalification['A24'] = "Type of Organization:"
    prequalification['B24'] = "Corporation"
    prequalification['C24'] = "Partnership"
    prequalification['D24'] = "Sole Proprietorship"
    prequalification['E24'] = "Other"
    
    company_detail_fields = [
        "Annual Revenue (Last 3 Years):", "Current Year:", "Previous Year:", "Two Years Prior:",
        "Number of Employees:"
    ]
    
    for i, field in enumerate(company_detail_fields, 25):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i in [26, 27, 28]:
            prequalification[f'B{i}'] = "$"
    
    # Licenses & Certifications
    license_fields = [
        "Contractor License Number:", "License Classification:", "State:", "Expiration Date:",
        "Minority Business Enterprise (MBE)?", "Women Business Enterprise (WBE)?",
        "Small Business Enterprise (SBE)?", "Other Certifications:"
    ]
    
    for i, field in enumerate(license_fields, 32):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i in [35, 36, 37]:
            prequalification[f'B{i}'] = "Yes"
            prequalification[f'C{i}'] = "No"
    
    # Insurance Information
    insurance_fields = [
        "General Liability Insurance Carrier:", "Policy Number:", "Expiration Date:",
        "Coverage Limits:", "Professional Liability Insurance?", "Workers Compensation Insurance?",
        "Bonding Capacity:", "Bonding Company:"
    ]
    
    for i, field in enumerate(insurance_fields, 41):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i in [45, 46]:
            prequalification[f'B{i}'] = "Yes"
            prequalification[f'C{i}'] = "No"
        if i == 47:
            prequalification[f'B{i}'] = "$"
    
    # Experience
    experience_fields = [
        "Years in Business:", "Areas of Specialty:", "Geographic Areas Served:",
        "Largest Contract Completed:", "Year:"
    ]
    
    for i, field in enumerate(experience_fields, 51):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i == 54:
            prequalification[f'B{i}'] = "$"
    
    # References
    prequalification['A58'] = "Please list three recent projects completed"
    
    reference_fields = [
        "Project 1 Name:", "Contact Person:", "Phone Number:", "Contract Amount:", "Completion Date:",
        "Project 2 Name:", "Contact Person:", "Phone Number:", "Contract Amount:", "Completion Date:",
        "Project 3 Name:", "Contact Person:", "Phone Number:", "Contract Amount:", "Completion Date:"
    ]
    
    for i, field in enumerate(reference_fields, 59):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i in [62, 67, 72]:
            prequalification[f'B{i}'] = "$"
    
    # Safety
    safety_fields = [
        "EMR (Experience Modification Rate):", "OSHA Citations (Last 3 Years):",
        "Safety Program in Place?"
    ]
    
    for i, field in enumerate(safety_fields, 79):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i == 81:
            prequalification[f'B{i}'] = "Yes"
            prequalification[f'C{i}'] = "No"
    
    # Legal
    legal_fields = [
        "Litigation in Last 5 Years?", "If Yes, Please Explain:",
        "Bankruptcy in Last 7 Years?", "Failed to Complete a Contract?"
    ]
    
    for i, field in enumerate(legal_fields, 84):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
        if i in [84, 86, 87]:
            prequalification[f'B{i}'] = "Yes"
            prequalification[f'C{i}'] = "No"
    
    # Signature
    prequalification['A90'] = "I certify that the information provided above is true and accurate."
    
    signature_fields = ["Name:", "Title:", "Signature:", "Date:"]
    
    for i, field in enumerate(signature_fields, 91):
        prequalification[f'A{i}'] = field
        prequalification[f'A{i}'].font = normal_font
    
    # ===== PROJECT BID PACKAGE =====
    bid_package = wb.create_sheet("Project Bid Package")
    
    # Set column widths
    bid_package.column_dimensions['A'].width = 25
    bid_package.column_dimensions['B'].width = 30
    bid_package.column_dimensions['C'].width = 15
    bid_package.column_dimensions['D'].width = 15
    bid_package.column_dimensions['E'].width = 15
    bid_package.column_dimensions['F'].width = 25
    
    # Title
    bid_package['A1'] = "PROJECT BID PACKAGE"
    bid_package['A1'].font = Font(name='Arial', size=14, bold=True)
    bid_package.merge_cells('A1:F1')
    bid_package['A1'].alignment = Alignment(horizontal='center')
    
    # Project Information
    bid_package['A3'] = "PROJECT INFORMATION"
    bid_package['A3'].font = header_font
    bid_package['A3'].fill = header_fill
    bid_package.merge_cells('A3:F3')
    
    project_info_fields = [
        "Project Name:", "Project Location:", "Owner:", "Architect/Engineer:",
        "Bid Package Number:", "Date Issued:", "Bid Due Date:"
    ]
    
    for i, field in enumerate(project_info_fields, 4):
        bid_package[f'A{i}'] = field
        bid_package[f'A{i}'].font = normal_font
    
    # Bid Package Scope
    bid_package['A12'] = "BID PACKAGE SCOPE"
    bid_package['A12'].font = header_font
    bid_package['A12'].fill = header_fill
    bid_package.merge_cells('A12:F12')
    
    # Header row
    scope_headers = ["Division", "Description", "Included", "Not Included", "Unit", "Notes"]
    for col, header in enumerate(scope_headers, 1):
        col_letter = get_column_letter(col)
        bid_package[f'{col_letter}13'] = header
        bid_package[f'{col_letter}13'].font = subheader_font
        bid_package[f'{col_letter}13'].fill = section_fill
    
    # CSI Divisions
    divisions = [
        ("Division 01", "General Requirements"),
        ("Division 02", "Existing Conditions"),
        ("Division 03", "Concrete"),
        ("Division 04", "Masonry"),
        ("Division 05", "Metals"),
        ("Division 06", "Wood, Plastics & Composites"),
        ("Division 07", "Thermal & Moisture Protection"),
        ("Division 08", "Openings"),
        ("Division 09", "Finishes"),
        ("Division 10", "Specialties"),
        ("Division 11", "Equipment"),
        ("Division 12", "Furnishings"),
        ("Division 13", "Special Construction"),
        ("Division 14", "Conveying Equipment"),
        ("Division 21", "Fire Suppression"),
        ("Division 22", "Plumbing"),
        ("Division 23", "HVAC"),
        ("Division 26", "Electrical"),
        ("Division 27", "Communications"),
        ("Division 28", "Electronic Safety & Security"),
        ("Division 31", "Earthwork"),
        ("Division 32", "Exterior Improvements"),
        ("Division 33", "Utilities")
    ]
    
    for i, (div_num, desc) in enumerate(divisions, 14):
        bid_package[f'A{i}'] = div_num
        bid_package[f'B{i}'] = desc
        bid_package[f'A{i}'].font = normal_font
        bid_package[f'B{i}'].font = normal_font
    
    # Schedule
    bid_package['A38'] = "SCHEDULE"
    bid_package['A38'].font = header_font
    bid_package['A38'].fill = header_fill
    bid_package.merge_cells('A38:F38')
    
    # Schedule Headers
    schedule_headers = ["Milestone", "Date", "Duration", "Notes", "", ""]
    for col, header in enumerate(schedule_headers, 1):
        col_letter = get_column_letter(col)
        bid_package[f'{col_letter}39'] = header
        bid_package[f'{col_letter}39'].font = subheader_font
        bid_package[f'{col_letter}39'].fill = section_fill
    
    schedule_milestones = [
        "Notice to Proceed", "Site Mobilization", "Substantial Completion", "Final Completion"
    ]
    
    for i, milestone in enumerate(schedule_milestones, 40):
        bid_package[f'A{i}'] = milestone
        bid_package[f'A{i}'].font = normal_font
    
    # Special Requirements
    bid_package['A45'] = "SPECIAL REQUIREMENTS"
    bid_package['A45'].font = header_font
    bid_package['A45'].fill = header_fill
    bid_package.merge_cells('A45:F45')
    
    # Special Requirements Headers
    special_req_headers = ["Item", "Description", "Required", "Notes", "", ""]
    for col, header in enumerate(special_req_headers, 1):
        col_letter = get_column_letter(col)
        bid_package[f'{col_letter}46'] = header
        bid_package[f'{col_letter}46'].font = subheader_font
        bid_package[f'{col_letter}46'].fill = section_fill
    
    special_req_items = [
        "Performance Bond", "Payment Bond", "Bid Bond", "Insurance Requirements",
        "Warranty Period", "Liquidated Damages", "Retainage"
    ]
    
    for i, item in enumerate(special_req_items, 47):
        bid_package[f'A{i}'] = item
        bid_package[f'A{i}'].font = normal_font
    
    # Attachments
    bid_package['A55'] = "ATTACHMENTS"
    bid_package['A55'].font = header_font
    bid_package['A55'].fill = header_fill
    bid_package.merge_cells('A55:F55')
    
    # Attachments Headers
    attachment_headers = ["Document", "Filename", "Version", "Date", "", ""]
    for col, header in enumerate(attachment_headers, 1):
        col_letter = get_column_letter(col)
        bid_package[f'{col_letter}56'] = header
        bid_package[f'{col_letter}56'].font = subheader_font
        bid_package[f'{col_letter}56'].fill = section_fill
    
    attachment_docs = [
        "Plans", "Specifications", "Geotechnical Report", "Survey", "Permits"
    ]
    
    for i, doc in enumerate(attachment_docs, 57):
        bid_package[f'A{i}'] = doc
        bid_package[f'A{i}'].font = normal_font
    
    # Addenda
    bid_package['A63'] = "ADDENDA"
    bid_package['A63'].font = header_font
    bid_package['A63'].fill = header_fill
    bid_package.merge_cells('A63:F63')
    
    # Addenda Headers
    addenda_headers = ["Addendum No.", "Issue Date", "Description", "", "", ""]
    for col, header in enumerate(addenda_headers, 1):
        col_letter = get_column_letter(col)
        bid_package[f'{col_letter}64'] = header
        bid_package[f'{col_letter}64'].font = subheader_font
        bid_package[f'{col_letter}64'].fill = section_fill
    
    # Questions & Clarifications
    bid_package['A70'] = "QUESTIONS & CLARIFICATIONS"
    bid_package['A70'].font = header_font
    bid_package['A70'].fill = header_fill
    bid_package.merge_cells('A70:F70')
    
    # Q&C Headers
    qc_headers = ["Question", "Response", "Date", "By", "", ""]
    for col, header in enumerate(qc_headers, 1):
        col_letter = get_column_letter(col)
        bid_package[f'{col_letter}71'] = header
        bid_package[f'{col_letter}71'].font = subheader_font
        bid_package[f'{col_letter}71'].fill = section_fill
    
    # ===== BID FORM =====
    bid_form = wb.create_sheet("Bid Form")
    
    # Set column widths
    bid_form.column_dimensions['A'].width = 25
    bid_form.column_dimensions['B'].width = 30
    bid_form.column_dimensions['C'].width = 20
    bid_form.column_dimensions['D'].width = 20
    bid_form.column_dimensions['E'].width = 20
    
    # Title
    bid_form['A1'] = "BID FORM"
    bid_form['A1'].font = Font(name='Arial', size=14, bold=True)
    bid_form.merge_cells('A1:E1')
    bid_form['A1'].alignment = Alignment(horizontal='center')
    
    # Project Information
    project_fields = [
        "Project Name:", "Bid Package Number:", "Contractor Name:", "Date:"
    ]
    
    for i, field in enumerate(project_fields, 3):
        bid_form[f'A{i}'] = field
        bid_form[f'A{i}'].font = normal_font
    
    # Base Bid
    bid_form['A8'] = "BASE BID"
    bid_form['A8'].font = header_font
    bid_form['A8'].fill = header_fill
    bid_form.merge_cells('A8:E8')
    
    bid_form['A9'] = "The undersigned Bidder, having examined the Bidding Documents and all other related documents, and being familiar with the site and conditions of the proposed work, hereby proposes to furnish all material, labor, tools, equipment, and supervision necessary to complete the work in accordance with the Bidding Documents for the following Base Bid amount:"
    bid_form['A9'].font = normal_font
    bid_form.merge_cells('A9:E9')
    bid_form['A9'].alignment = Alignment(wrap_text=True)
    
    bid_form['A11'] = "Base Bid Amount: $"
    bid_form['A12'] = "Base Bid (Written):"
    bid_form['A11'].font = normal_font
    bid_form['A12'].font = normal_font
    
    # Alternates
    bid_form['A14'] = "ALTERNATES"
    bid_form['A14'].font = header_font
    bid_form['A14'].fill = header_fill
    bid_form.merge_cells('A14:E14')
    
    # Alternates Headers
    alternates_headers = ["Alternate No.", "Description", "ADD", "DEDUCT", "Calendar Days"]
    for col, header in enumerate(alternates_headers, 1):
        col_letter = get_column_letter(col)
        bid_form[f'{col_letter}15'] = header
        bid_form[f'{col_letter}15'].font = subheader_font
        bid_form[f'{col_letter}15'].fill = section_fill
    
    alternates = ["Alternate 1", "Alternate 2", "Alternate 3"]
    
    for i, alt in enumerate(alternates, 16):
        bid_form[f'A{i}'] = alt
        bid_form[f'A{i}'].font = normal_font
        bid_form[f'C{i}'] = "$"
        bid_form[f'D{i}'] = "$"
    
    # Unit Prices
    bid_form['A20'] = "UNIT PRICES"
    bid_form['A20'].font = header_font
    bid_form['A20'].fill = header_fill
    bid_form.merge_cells('A20:E20')
    
    # Unit Prices Headers
    unit_price_headers = ["Unit Price No.", "Description", "Unit", "Price", ""]
    for col, header in enumerate(unit_price_headers, 1):
        col_letter = get_column_letter(col)
        bid_form[f'{col_letter}21'] = header
        bid_form[f'{col_letter}21'].font = subheader_font
        bid_form[f'{col_letter}21'].fill = section_fill
    
    unit_prices = ["Unit Price 1", "Unit Price 2", "Unit Price 3"]
    
    for i, up in enumerate(unit_prices, 22):
        bid_form[f'A{i}'] = up
        bid_form[f'A{i}'].font = normal_font
        bid_form[f'D{i}'] = "$"
    
    # Allowances
    bid_form['A26'] = "ALLOWANCES"
    bid_form['A26'].font = header_font
    bid_form['A26'].fill = header_fill
    bid_form.merge_cells('A26:E26')
    
    # Allowances Headers
    allowance_headers = ["Allowance No.", "Description", "Amount", "", ""]
    for col, header in enumerate(allowance_headers, 1):
        col_letter = get_column_letter(col)
        bid_form[f'{col_letter}27'] = header
        bid_form[f'{col_letter}27'].font = subheader_font
        bid_form[f'{col_letter}27'].fill = section_fill
    
    allowances = ["Allowance 1", "Allowance 2", "Allowance 3"]
    
    for i, allow in enumerate(allowances, 28):
        bid_form[f'A{i}'] = allow
        bid_form[f'A{i}'].font = normal_font
        bid_form[f'C{i}'] = "$"
    
    # Time of Completion
    bid_form['A32'] = "TIME OF COMPLETION"
    bid_form['A32'].font = header_font
    bid_form['A32'].fill = header_fill
    bid_form.merge_cells('A32:E32')
    
    bid_form['A33'] = "The undersigned Bidder proposes to achieve Substantial Completion of the Work within"
    bid_form['C33'] = "calendar days from the date of Notice to Proceed."
    bid_form['A33'].font = normal_font
    bid_form['C33'].font = normal_font
    bid_form.merge_cells('A33:B33')
    bid_form.merge_cells('C33:E33')
    
    # Addenda
    bid_form['A35'] = "ADDENDA"
    bid_form['A35'].font = header_font
    bid_form['A35'].fill = header_fill
    bid_form.merge_cells('A35:E35')
    
    bid_form['A36'] = "The Bidder acknowledges receipt of the following Addenda:"
    bid_form['A36'].font = normal_font
    bid_form.merge_cells('A36:E36')
    
    bid_form['A37'] = "Addendum No."
    bid_form['B37'] = "Date"
    bid_form['A37'].font = subheader_font
    bid_form['B37'].font = subheader_font
    bid_form['A37'].fill = section_fill
    bid_form['B37'].fill = section_fill
    
    # Bid Security
    bid_form['A43'] = "BID SECURITY"
    bid_form['A43'].font = header_font
    bid_form['A43'].fill = header_fill
    bid_form.merge_cells('A43:E43')
    
    bid_form['A44'] = "Bid Security in the amount of 5% of the Base Bid is enclosed in the form of:"
    bid_form['A44'].font = normal_font
    bid_form.merge_cells('A44:E44')
    
    bid_form['A45'] = "☐ Bid Bond"
    bid_form['B45'] = "☐ Certified Check"
    bid_form['C45'] = "☐ Cashier's Check"
    bid_form['A45'].font = normal_font
    bid_form['B45'].font = normal_font
    bid_form['C45'].font = normal_font
    
    # Subcontractors
    bid_form['A47'] = "SUBCONTRACTORS"
    bid_form['A47'].font = header_font
    bid_form['A47'].fill = header_fill
    bid_form.merge_cells('A47:E47')
    
    bid_form['A48'] = "The Bidder proposes to use the following subcontractors for the portions of work indicated:"
    bid_form['A48'].font = normal_font
    bid_form.merge_cells('A48:E48')
    
    bid_form['A49'] = "Portion of Work"
    bid_form['B49'] = "Subcontractor Name"
    bid_form['A49'].font = subheader_font
    bid_form['B49'].font = subheader_font
    bid_form['A49'].fill = section_fill
    bid_form['B49'].fill = section_fill
    
    # Bidder Information
    bid_form['A54'] = "BIDDER INFORMATION"
    bid_form['A54'].font = header_font
    bid_form['A54'].fill = header_fill
    bid_form.merge_cells('A54:E54')
    
    bidder_info_fields = [
        "Company Name:", "Address:", "City, State, ZIP:", "Phone:", "Email:",
        "Contractor License No.:"
    ]
    
    for i, field in enumerate(bidder_info_fields, 55):
        bid_form[f'A{i}'] = field
        bid_form[f'A{i}'].font = normal_font
    
    # Signature
    bid_form['A62'] = "SIGNATURE"
    bid_form['A62'].font = header_font
    bid_form['A62'].fill = header_fill
    bid_form.merge_cells('A62:E62')
    
    bid_form['A63'] = "The undersigned hereby certifies that the Bidder has examined and fully understands the requirements and conditions of the Bidding Documents, has examined the site and all conditions affecting the Work, and proposes to provide all required labor, materials, and equipment to perform the Work in strict accordance with the Bidding Documents."
    bid_form['A63'].font = normal_font
    bid_form.merge_cells('A63:E63')
    bid_form['A63'].alignment = Alignment(wrap_text=True)
    
    signature_fields = ["Signature:", "Name (Printed):", "Title:", "Date:"]
    
    for i, field in enumerate(signature_fields, 65):
        bid_form[f'A{i}'] = field
        bid_form[f'A{i}'].font = normal_font
    
    # ===== INSTRUCTIONS TO BIDDERS =====
    instructions = wb.create_sheet("Instructions to Bidders")
    
    # Set column widths
    instructions.column_dimensions['A'].width = 15
    instructions.column_dimensions['B'].width = 70
    instructions.column_dimensions['C'].width = 15
    instructions.column_dimensions['D'].width = 15
    instructions.column_dimensions['E'].width = 15
    
    # Title
    instructions['A1'] = "INSTRUCTIONS TO BIDDERS"
    instructions['A1'].font = Font(name='Arial', size=14, bold=True)
    instructions.merge_cells('A1:E1')
    instructions['A1'].alignment = Alignment(horizontal='center')
    
    # Project Information
    instructions['A3'] = "PROJECT INFORMATION"
    instructions['A3'].font = header_font
    instructions['A3'].fill = header_fill
    instructions.merge_cells('A3:E3')
    
    project_info_fields = [
        "Project Name:", "Project Location:", "Owner:", "Architect/Engineer:"
    ]
    
    for i, field in enumerate(project_info_fields, 4):
        instructions[f'A{i}'] = field
        instructions[f'A{i}'].font = normal_font
    
    # Bidding Requirements
    instructions['A9'] = "BIDDING REQUIREMENTS"
    instructions['A9'].font = header_font
    instructions['A9'].fill = header_fill
    instructions.merge_cells('A9:E9')
    
    # Bidding Requirements Headers
    bid_req_headers = ["Item", "Description", "Due Date", "Notes", ""]
    for col, header in enumerate(bid_req_headers, 1):
        col_letter = get_column_letter(col)
        instructions[f'{col_letter}10'] = header
        instructions[f'{col_letter}10'].font = subheader_font
        instructions[f'{col_letter}10'].fill = section_fill
    
    bid_req_items = [
        "Mandatory Pre-Bid Meeting", "Questions Deadline", 
        "Bid Submission Deadline", "Bid Opening"
    ]
    
    for i, item in enumerate(bid_req_items, 11):
        instructions[f'A{i}'] = item
        instructions[f'A{i}'].font = normal_font
    
    # Sections of Instructions
    sections = [
        (16, "1. DEFINITIONS"),
        (22, "2. BIDDER'S REPRESENTATION"),
        (28, "3. BIDDING DOCUMENTS"),
        (34, "4. INTERPRETATIONS AND ADDENDA"),
        (40, "5. BIDDING PROCEDURES"),
        (48, "6. BID SECURITY"),
        (54, "7. SUBMISSION OF BIDS"),
        (60, "8. MODIFICATION AND WITHDRAWAL OF BIDS"),
        (64, "9. OPENING OF BIDS"),
        (68, "10. AWARD OF CONTRACT"),
        (74, "11. SUBCONTRACTORS"),
        (78, "12. BONDS AND INSURANCE"),
        (83, "13. TIME OF COMPLETION"),
        (87, "14. APPLICABLE LAWS"),
        (91, "15. POST-BID INFORMATION")
    ]
    
    for row, title in sections:
        instructions[f'A{row}'] = title
        instructions[f'A{row}'].font = header_font
        instructions[f'A{row}'].fill = header_fill
        instructions.merge_cells(f'A{row}:E{row}')
    
    # Definitions Section
    definitions = [
        ("1.1", "Bidding Documents: Contract Documents including Invitation to Bid, Instructions to Bidders, Bid Form, and proposed Contract Documents including Drawings and Specifications."),
        ("1.2", "Addenda: Written or graphic changes or interpretations of the Contract Documents issued prior to the bid opening."),
        ("1.3", "Base Bid: The sum of money stated in the Bid for which the Bidder offers to perform the Work."),
        ("1.4", "Alternate: An amount stated in the Bid to be added to or deducted from the Base Bid if the corresponding change in the Work is accepted."),
        ("1.5", "Unit Price: An amount stated in the Bid as a price per unit of measurement for materials, equipment, or services.")
    ]
    
    for i, (num, desc) in enumerate(definitions, 17):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Bidder's Representation Section
    instructions['A23'] = "2.1"
    instructions['B23'] = "By submitting a bid, the Bidder represents that:"
    
    representations = [
        ("2.1.1", "The Bidder has read and understands the Bidding Documents."),
        ("2.1.2", "The Bidder has visited the site and is familiar with local conditions under which the Work is to be performed."),
        ("2.1.3", "The Bid is based upon the materials, equipment, and systems required by the Bidding Documents."),
        ("2.1.4", "The Bidder has the capability, experience, and resources to perform the Work as specified.")
    ]
    
    for i, (num, desc) in enumerate(representations, 24):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Bidding Documents Section
    bidding_docs = [
        ("3.1", "Copies of the Bidding Documents may be obtained from:"),
        ("3.2", "Bidding Documents will be available for inspection at:"),
        ("3.3", "Bidders shall use complete sets of Bidding Documents in preparing Bids."),
        ("3.4", "Bidders shall promptly notify the Owner of any inconsistencies or errors discovered in the Bidding Documents.")
    ]
    
    for i, (num, desc) in enumerate(bidding_docs, 29):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Interpretations and Addenda Section
    interpretations = [
        ("4.1", "Questions regarding the Bidding Documents shall be submitted in writing to:"),
        ("4.2", "Interpretations, corrections, and changes will be made by Addenda sent to all Bidders."),
        ("4.3", "Addenda will be issued no later than [X] days prior to the bid opening date."),
        ("4.4", "Each Bidder shall acknowledge receipt of all Addenda on the Bid Form.")
    ]
    
    for i, (num, desc) in enumerate(interpretations, 35):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Bidding Procedures Section
    procedures = [
        ("5.1", "Bids shall be submitted on the Bid Form provided in the Bidding Documents."),
        ("5.2", "All blanks on the Bid Form shall be completed in ink or typed."),
        ("5.3", "Alternates shall be bid as requested on the Bid Form."),
        ("5.4", "Unit Prices shall be shown on the Bid Form where required."),
        ("5.5", "Bids shall not contain any conditions or qualifications not provided for in the Bid Form."),
        ("5.6", "Bids shall be signed by an authorized representative of the Bidder.")
    ]
    
    for i, (num, desc) in enumerate(procedures, 41):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Bid Security Section
    bid_security = [
        ("6.1", "Each Bid shall be accompanied by a Bid Security in the amount of [X]% of the Base Bid."),
        ("6.2", "The Bid Security shall be in the form of a Bid Bond, Certified Check, or Cashier's Check."),
        ("6.3", "The Bid Security of the successful Bidder will be retained until the Contract has been executed."),
        ("6.4", "The Bid Security of unsuccessful Bidders will be returned upon award of the Contract.")
    ]
    
    for i, (num, desc) in enumerate(bid_security, 49):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Submission of Bids Section
    submission = [
        ("7.1", "Bids shall be submitted in a sealed envelope with the following information clearly marked on the outside:"),
        ("7.1.1", "Project Name"),
        ("7.1.2", "Bid Package Number"),
        ("7.1.3", "Bidder's Name and Address"),
        ("7.2", "Bids shall be delivered to:"),
        ("7.3", "Bids must be received by the date and time specified in the Invitation to Bid."),
        ("7.4", "Late Bids will not be accepted.")
    ]
    
    for i, (num, desc) in enumerate(submission, 55):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Modification and Withdrawal Section
    modification = [
        ("8.1", "Bids may be modified or withdrawn by written notice received prior to the bid opening."),
        ("8.2", "No Bid may be withdrawn for a period of [X] days after the bid opening without written consent of the Owner.")
    ]
    
    for i, (num, desc) in enumerate(modification, 61):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Opening of Bids Section
    opening = [
        ("9.1", "Bids will be opened publicly at the time and place specified in the Invitation to Bid."),
        ("9.2", "The Owner reserves the right to reject any or all Bids and to waive informalities.")
    ]
    
    for i, (num, desc) in enumerate(opening, 65):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Award of Contract Section
    award = [
        ("10.1", "The Contract will be awarded to the lowest responsive, responsible Bidder, if awarded."),
        ("10.2", "The Owner reserves the right to accept or reject Alternates in any order or combination."),
        ("10.3", "The successful Bidder will be required to execute the Contract within [X] days after notification of award."),
        ("10.4", "The successful Bidder will be required to furnish Performance and Payment Bonds.")
    ]
    
    for i, (num, desc) in enumerate(award, 69):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Subcontractors Section
    subcontractors = [
        ("11.1", "The Bidder shall list on the Bid Form all major Subcontractors proposed for portions of the Work."),
        ("11.2", "The Owner reserves the right to reject any proposed Subcontractor.")
    ]
    
    for i, (num, desc) in enumerate(subcontractors, 75):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Bonds and Insurance Section
    bonds = [
        ("12.1", "The successful Bidder shall furnish Performance and Payment Bonds in the amount of 100% of the Contract Sum."),
        ("12.2", "The cost of such bonds shall be included in the Base Bid."),
        ("12.3", "The successful Bidder shall provide Certificates of Insurance as required by the Contract Documents.")
    ]
    
    for i, (num, desc) in enumerate(bonds, 79):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Time of Completion Section
    completion = [
        ("13.1", "The Work shall be commenced and completed as specified in the Contract Documents."),
        ("13.2", "Liquidated damages for delay may be assessed as specified in the Contract Documents.")
    ]
    
    for i, (num, desc) in enumerate(completion, 84):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Applicable Laws Section
    laws = [
        ("14.1", "The Bidder shall comply with all applicable federal, state, and local laws and regulations."),
        ("14.2", "The Bidder shall pay all taxes, fees, and assessments as required.")
    ]
    
    for i, (num, desc) in enumerate(laws, 88):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # Post-Bid Information Section
    post_bid = [
        ("15.1", "The successful Bidder shall submit the following information within [X] days of the bid opening:"),
        ("15.1.1", "A complete list of Subcontractors and Suppliers."),
        ("15.1.2", "A Schedule of Values."),
        ("15.1.3", "A Construction Schedule.")
    ]
    
    for i, (num, desc) in enumerate(post_bid, 92):
        instructions[f'A{i}'] = num
        instructions[f'B{i}'] = desc
        instructions[f'A{i}'].font = normal_font
        instructions[f'B{i}'].font = normal_font
    
    # ===== BID MANUAL TABLE OF CONTENTS =====
    toc = wb.create_sheet("Bid Manual TOC")
    
    # Set column widths
    toc.column_dimensions['A'].width = 10
    toc.column_dimensions['B'].width = 50
    toc.column_dimensions['C'].width = 10
    
    # Title
    toc['A1'] = "BID MANUAL - TABLE OF CONTENTS"
    toc['A1'].font = Font(name='Arial', size=14, bold=True)
    toc.merge_cells('A1:C1')
    toc['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    toc['A3'] = "SECTION"
    toc['B3'] = "DESCRIPTION"
    toc['C3'] = "PAGE"
    toc['A3'].font = header_font
    toc['B3'].font = header_font
    toc['C3'].font = header_font
    toc['A3'].fill = header_fill
    toc['B3'].fill = header_fill
    toc['C3'].fill = header_fill
    
    # TOC Items
    toc_items = [
        ("1", "Project Information"),
        ("2", "Invitation to Bid"),
        ("3", "Instructions to Bidders"),
        ("4", "Bid Form"),
        ("5", "Bid Bond Form"),
        ("6", "Contract Form"),
        ("7", "General Conditions"),
        ("8", "Supplementary Conditions"),
        ("9", "Technical Specifications"),
        ("10", "Drawings List"),
        ("11", "Addenda"),
        ("12", "Prevailing Wage Rates (if applicable)"),
        ("13", "Insurance Requirements"),
        ("14", "Sample Forms"),
        ("15", "Project Schedule")
    ]
    
    for i, (num, desc) in enumerate(toc_items, 4):
        toc[f'A{i}'] = num
        toc[f'B{i}'] = desc
        toc[f'A{i}'].font = normal_font
        toc[f'B{i}'].font = normal_font
    
    # Apply some additional styling to improve appearance
    def apply_borders_to_sheet(sheet, start_row, end_row, start_col, end_col):
        """Apply borders to a range of cells in a sheet"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
    
    # Apply borders to key sections of each sheet
    apply_borders_to_sheet(prequalification, 4, 92, 1, 5)  # Prequalification Form
    apply_borders_to_sheet(bid_package, 4, 75, 1, 6)  # Bid Package
    apply_borders_to_sheet(bid_form, 3, 69, 1, 5)  # Bid Form
    apply_borders_to_sheet(instructions, 4, 96, 1, 5)  # Instructions
    apply_borders_to_sheet(toc, 3, 18, 1, 3)  # Table of Contents
    
    # Set the Prequalification Form as the active sheet when the workbook is opened
    wb.active = 0
    
    # Save the workbook
    wb.save(filename)
    print(f"Construction Bidding Workbook created: {filename}")
    return filename

# Run the function to create the workbook
if __name__ == "__main__":
    create_construction_bidding_workbook()
