import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import openpyxl
from datetime import datetime

def create_construction_budget_workbook(filename='Construction_Budget_Workbook.xlsx'):
    """
    Creates a comprehensive Excel workbook for construction budget management
    with multiple sheets for budget plan, forecasting, and cost tracking.
    """
    # Create a new workbook
    wb = Workbook()
    
    # Set workbook properties
    wb.properties.creator = 'Construction Budget Tool'
    wb.properties.lastModifiedBy = 'Project Manager'
    wb.properties.created = datetime.now()
    wb.properties.modified = datetime.now()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create all sheets
    budget_sheet = wb.create_sheet("Budget Plan")
    forecast_sheet = wb.create_sheet("Budget Forecast")
    cost_tracking_sheet = wb.create_sheet("Cost Tracking")
    dashboard_sheet = wb.create_sheet("Dashboard")
    instructions_sheet = wb.create_sheet("Instructions")
    
    # Set tab colors
    budget_sheet.sheet_properties.tabColor = "4F81BD"
    forecast_sheet.sheet_properties.tabColor = "9BBB59"
    cost_tracking_sheet.sheet_properties.tabColor = "C0504D"
    dashboard_sheet.sheet_properties.tabColor = "4BACC6"
    instructions_sheet.sheet_properties.tabColor = "808080"
    
    # Define common styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    forecast_header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    cost_tracking_header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    dashboard_header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create Budget Plan Sheet
    create_budget_plan_sheet(wb, budget_sheet, header_font, header_fill, alt_row_fill, thin_border)
    
    # Create Budget Forecast Sheet
    create_budget_forecast_sheet(wb, forecast_sheet, header_font, forecast_header_fill, alt_row_fill, thin_border)
    
    # Create Cost Tracking Sheet
    create_cost_tracking_sheet(wb, cost_tracking_sheet, header_font, cost_tracking_header_fill, alt_row_fill, thin_border)
    
    # Create Dashboard Sheet
    create_dashboard_sheet(wb, dashboard_sheet, header_font, dashboard_header_fill, thin_border)
    
    # Create Instructions Sheet
    create_instructions_sheet(wb, instructions_sheet, header_font)
    
    # Save the workbook
    wb.save(filename)
    print(f"Construction Budget Workbook created successfully as {filename}")

def create_budget_plan_sheet(wb, sheet, header_font, header_fill, alt_row_fill, thin_border):
    """Creates and formats the Budget Plan sheet"""
    
    # Set column widths
    columns = [
        ('Category', 20), ('Subcategory', 25), ('Description', 30), 
        ('Estimated Cost', 15), ('Unit', 10), ('Quantity', 10), 
        ('Unit Price', 12), ('Total Budget', 15), ('Notes', 30)
    ]
    
    for i, (header, width) in enumerate(columns, start=1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width
        cell = sheet.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Create budget categories data
    budget_categories = [
        ('Pre-Construction', 'Site Survey', 'Topographical survey of site', 'Acres', 5, 1000),
        ('Pre-Construction', 'Permits', 'Building permits', 'Lump Sum', 1, 15000),
        ('Pre-Construction', 'Architecture/Engineering', 'Design services', 'Lump Sum', 1, 75000),
        ('Site Work', 'Excavation', 'Site preparation and grading', 'Cu. Yards', 500, 30),
        ('Site Work', 'Utilities', 'Water, sewer, electric connections', 'Lump Sum', 1, 35000),
        ('Foundation', 'Concrete', 'Foundation concrete', 'Cu. Yards', 120, 200),
        ('Foundation', 'Waterproofing', 'Foundation waterproofing', 'Sq. Ft.', 2000, 5),
        ('Framing', 'Wood Framing', 'Structural framing materials', 'Sq. Ft.', 3500, 25),
        ('Framing', 'Labor', 'Framing labor', 'Hours', 500, 50),
        ('Exterior', 'Roofing', 'Roofing materials and installation', 'Sq. Ft.', 2000, 15),
        ('Exterior', 'Siding', 'Exterior siding', 'Sq. Ft.', 3000, 12),
        ('Exterior', 'Windows & Doors', 'Exterior doors and windows', 'Each', 25, 1200),
        ('Interior', 'Drywall', 'Drywall installation', 'Sq. Ft.', 5000, 4),
        ('Interior', 'Flooring', 'Various flooring materials', 'Sq. Ft.', 3500, 12),
        ('Interior', 'Painting', 'Interior painting', 'Sq. Ft.', 5000, 3),
        ('Interior', 'Trim & Doors', 'Interior trim and doors', 'Lump Sum', 1, 22000),
        ('MEP', 'Electrical', 'Electrical systems', 'Lump Sum', 1, 45000),
        ('MEP', 'Plumbing', 'Plumbing systems', 'Lump Sum', 1, 38000),
        ('MEP', 'HVAC', 'Heating and cooling systems', 'Lump Sum', 1, 42000),
        ('Fixtures', 'Kitchen', 'Kitchen cabinets and countertops', 'Lump Sum', 1, 30000),
        ('Fixtures', 'Bathroom', 'Bathroom fixtures and tile', 'Each', 3, 12000),
        ('Fixtures', 'Appliances', 'Kitchen appliances', 'Lump Sum', 1, 15000),
        ('Landscaping', 'Grading', 'Final grading', 'Sq. Ft.', 10000, 0.75),
        ('Landscaping', 'Planting', 'Trees, shrubs, and lawn', 'Lump Sum', 1, 15000),
        ('Other', 'Cleaning', 'Final cleaning', 'Lump Sum', 1, 3500),
        ('Other', 'Dumpsters', 'Waste removal', 'Each', 10, 600),
    ]
    
    # Add data rows
    row = 2  # Start from row 2 after headers
    for category, subcategory, description, unit, quantity, unit_price in budget_categories:
        # Calculate the total
        total = quantity * unit_price
        
        # Add data to cells
        sheet.cell(row=row, column=1, value=category).border = thin_border
        sheet.cell(row=row, column=2, value=subcategory).border = thin_border
        sheet.cell(row=row, column=3, value=description).border = thin_border
        
        # For Estimated Cost, use formula (will update later)
        est_cost_cell = sheet.cell(row=row, column=4)
        est_cost_cell.value = f"=G{row}*H{row}"
        est_cost_cell.number_format = '$#,##0.00'
        est_cost_cell.border = thin_border
        
        sheet.cell(row=row, column=5, value=unit).border = thin_border
        
        qty_cell = sheet.cell(row=row, column=6, value=quantity)
        qty_cell.number_format = '#,##0.00'
        qty_cell.border = thin_border
        
        unit_price_cell = sheet.cell(row=row, column=7, value=unit_price)
        unit_price_cell.number_format = '$#,##0.00'
        unit_price_cell.border = thin_border
        
        total_cell = sheet.cell(row=row, column=8)
        total_cell.value = f"=D{row}"
        total_cell.number_format = '$#,##0.00'
        total_cell.border = thin_border
        
        notes_cell = sheet.cell(row=row, column=9, value="")
        notes_cell.border = thin_border
        
        # Apply alternating row formatting
        if row % 2 == 0:
            for col in range(1, 10):
                sheet.cell(row=row, column=col).fill = alt_row_fill
        
        row += 1
    
    # Add contingency row
    sheet.cell(row=row, column=1, value="Contingency").border = thin_border
    sheet.cell(row=row, column=2, value="Contingency Fund").border = thin_border
    sheet.cell(row=row, column=3, value="10% of total budget").border = thin_border
    
    est_cost_cell = sheet.cell(row=row, column=4)
    est_cost_cell.value = f"=G{row}*H{row}"
    est_cost_cell.number_format = '$#,##0.00'
    est_cost_cell.border = thin_border
    
    sheet.cell(row=row, column=5, value="Percentage").border = thin_border
    
    contingency_cell = sheet.cell(row=row, column=6, value=0.1)
    contingency_cell.number_format = '#,##0.00'
    contingency_cell.border = thin_border
    
    # For unit price of contingency, sum all totals above
    unit_price_cell = sheet.cell(row=row, column=7)
    unit_price_cell.value = f"=SUM(H2:H{row-1})"
    unit_price_cell.number_format = '$#,##0.00'
    unit_price_cell.border = thin_border
    
    total_cell = sheet.cell(row=row, column=8)
    total_cell.value = f"=D{row}"
    total_cell.number_format = '$#,##0.00'
    total_cell.border = thin_border
    
    notes_cell = sheet.cell(row=row, column=9, value="For unexpected costs")
    notes_cell.border = thin_border
    
    if row % 2 == 0:
        for col in range(1, 10):
            sheet.cell(row=row, column=col).fill = alt_row_fill
    
    # Add total row
    row += 1
    sheet.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    sheet.cell(row=row, column=1).border = thin_border
    sheet.cell(row=row, column=2, value="").border = thin_border
    sheet.cell(row=row, column=3, value="").border = thin_border
    sheet.cell(row=row, column=4, value="").border = thin_border
    sheet.cell(row=row, column=5, value="").border = thin_border
    sheet.cell(row=row, column=6, value="").border = thin_border
    sheet.cell(row=row, column=7, value="").border = thin_border
    
    # Sum all totals including contingency
    total_sum_cell = sheet.cell(row=row, column=8)
    total_sum_cell.value = f"=SUM(H2:H{row-1})"
    total_sum_cell.number_format = '$#,##0.00'
    total_sum_cell.font = Font(bold=True)
    total_sum_cell.border = thin_border
    total_sum_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    sheet.cell(row=row, column=9, value="").border = thin_border

def create_budget_forecast_sheet(wb, sheet, header_font, header_fill, alt_row_fill, thin_border):
    """Creates and formats the Budget Forecast sheet"""
    
    # Set column widths
    columns = [
        ('Category', 20), ('Subcategory', 25), ('Total Budget', 15),
        ('Month 1', 12), ('Month 2', 12), ('Month 3', 12), ('Month 4', 12),
        ('Month 5', 12), ('Month 6', 12), ('Month 7', 12), ('Month 8', 12),
        ('Total Forecast', 15), ('Variance', 15)
    ]
    
    for i, (header, width) in enumerate(columns, start=1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width
        cell = sheet.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Link data from Budget Plan sheet
    budget_sheet = wb["Budget Plan"]
    rows = budget_sheet.max_row
    
    # For each row in Budget Plan, create a corresponding row in Forecast
    for i in range(2, rows): # Skip header row and total row
        row = i
        
        # Copy category and subcategory
        sheet.cell(row=row, column=1, value=f"='Budget Plan'!A{row}").border = thin_border
        sheet.cell(row=row, column=2, value=f"='Budget Plan'!B{row}").border = thin_border
        
        # Link to budget total
        budget_cell = sheet.cell(row=row, column=3, value=f"='Budget Plan'!H{row}")
        budget_cell.number_format = '$#,##0.00'
        budget_cell.border = thin_border
        
        # Initialize monthly forecast cells to 0
        for month in range(4, 12): # Months 1-8
            month_cell = sheet.cell(row=row, column=month, value=0)
            month_cell.number_format = '$#,##0.00'
            month_cell.border = thin_border
        
        # Set up formulas for Total Forecast and Variance
        total_forecast_cell = sheet.cell(row=row, column=12)
        total_forecast_cell.value = f"=SUM(D{row}:K{row})"
        total_forecast_cell.number_format = '$#,##0.00'
        total_forecast_cell.border = thin_border
        
        variance_cell = sheet.cell(row=row, column=13)
        variance_cell.value = f"=C{row}-L{row}"
        variance_cell.number_format = '$#,##0.00'
        variance_cell.border = thin_border
        
        # Apply alternating row formatting
        if row % 2 == 0:
            for col in range(1, 14):
                sheet.cell(row=row, column=col).fill = alt_row_fill
    
    # Add sample forecast distribution
    # Pre-Construction - mostly in month 1
    sheet.cell(row=2, column=4).value = "=C2 * 1.0"  # Site Survey - Month 1 (100%)
    sheet.cell(row=3, column=4).value = "=C3 * 1.0"  # Permits - Month 1 (100%)
    sheet.cell(row=4, column=4).value = "=C4 * 0.6"  # Architecture - Month 1 (60%)
    sheet.cell(row=4, column=5).value = "=C4 * 0.4"  # Architecture - Month 2 (40%)
    
    # Site Work - month 1-2
    sheet.cell(row=5, column=4).value = "=C5 * 0.7"  # Excavation - Month 1 (70%)
    sheet.cell(row=5, column=5).value = "=C5 * 0.3"  # Excavation - Month 2 (30%)
    sheet.cell(row=6, column=4).value = "=C6 * 0.3"  # Utilities - Month 1 (30%)
    sheet.cell(row=6, column=5).value = "=C6 * 0.7"  # Utilities - Month 2 (70%)
    
    # Foundation - month 2-3
    sheet.cell(row=7, column=5).value = "=C7 * 0.6"  # Concrete - Month 2 (60%)
    sheet.cell(row=7, column=6).value = "=C7 * 0.4"  # Concrete - Month 3 (40%)
    sheet.cell(row=8, column=5).value = "=C8 * 0.5"  # Waterproofing - Month 2 (50%)
    sheet.cell(row=8, column=6).value = "=C8 * 0.5"  # Waterproofing - Month 3 (50%)
    
    # Framing - month 3-4
    sheet.cell(row=9, column=6).value = "=C9 * 0.6"  # Wood Framing - Month 3 (60%)
    sheet.cell(row=9, column=7).value = "=C9 * 0.4"  # Wood Framing - Month 4 (40%)
    sheet.cell(row=10, column=6).value = "=C10 * 0.6"  # Framing Labor - Month 3 (60%)
    sheet.cell(row=10, column=7).value = "=C10 * 0.4"  # Framing Labor - Month 4 (40%)
    
    # Exterior - month 4-5
    sheet.cell(row=11, column=7).value = "=C11 * 0.8"  # Roofing - Month 4 (80%)
    sheet.cell(row=11, column=8).value = "=C11 * 0.2"  # Roofing - Month 5 (20%)
    sheet.cell(row=12, column=7).value = "=C12 * 0.3"  # Siding - Month 4 (30%)
    sheet.cell(row=12, column=8).value = "=C12 * 0.7"  # Siding - Month 5 (70%)
    sheet.cell(row=13, column=7).value = "=C13 * 0.4"  # Windows & Doors - Month 4 (40%)
    sheet.cell(row=13, column=8).value = "=C13 * 0.6"  # Windows & Doors - Month 5 (60%)
    
    # Interior - month 5-7
    sheet.cell(row=14, column=8).value = "=C14 * 0.7"  # Drywall - Month 5 (70%)
    sheet.cell(row=14, column=9).value = "=C14 * 0.3"  # Drywall - Month 6 (30%)
    sheet.cell(row=15, column=8).value = "=C15 * 0.2"  # Flooring - Month 5 (20%)
    sheet.cell(row=15, column=9).value = "=C15 * 0.6"  # Flooring - Month 6 (60%)
    sheet.cell(row=15, column=10).value = "=C15 * 0.2"  # Flooring - Month 7 (20%)
    sheet.cell(row=16, column=9).value = "=C16 * 0.7"  # Painting - Month 6 (70%)
    sheet.cell(row=16, column=10).value = "=C16 * 0.3"  # Painting - Month 7 (30%)
    sheet.cell(row=17, column=9).value = "=C17 * 0.6"  # Trim & Doors - Month 6 (60%)
    sheet.cell(row=17, column=10).value = "=C17 * 0.4"  # Trim & Doors - Month 7 (40%)
    
    # MEP - month 3-6
    sheet.cell(row=18, column=6).value = "=C18 * 0.2"  # Electrical - Month 3 (20%)
    sheet.cell(row=18, column=7).value = "=C18 * 0.3"  # Electrical - Month 4 (30%)
    sheet.cell(row=18, column=8).value = "=C18 * 0.3"  # Electrical - Month 5 (30%)
    sheet.cell(row=18, column=9).value = "=C18 * 0.2"  # Electrical - Month 6 (20%)
    sheet.cell(row=19, column=6).value = "=C19 * 0.2"  # Plumbing - Month 3 (20%)
    sheet.cell(row=19, column=7).value = "=C19 * 0.3"  # Plumbing - Month 4 (30%)
    sheet.cell(row=19, column=8).value = "=C19 * 0.3"  # Plumbing - Month 5 (30%)
    sheet.cell(row=19, column=9).value = "=C19 * 0.2"  # Plumbing - Month 6 (20%)
    sheet.cell(row=20, column=7).value = "=C20 * 0.3"  # HVAC - Month 4 (30%)
    sheet.cell(row=20, column=8).value = "=C20 * 0.4"  # HVAC - Month 5 (40%)
    sheet.cell(row=20, column=9).value = "=C20 * 0.3"  # HVAC - Month 6 (30%)
    
    # Fixtures - month 6-7
    sheet.cell(row=21, column=9).value = "=C21 * 0.7"  # Kitchen - Month 6 (70%)
    sheet.cell(row=21, column=10).value = "=C21 * 0.3"  # Kitchen - Month 7 (30%)
    sheet.cell(row=22, column=9).value = "=C22 * 0.5"  # Bathroom - Month 6 (50%)
    sheet.cell(row=22, column=10).value = "=C22 * 0.5"  # Bathroom - Month 7 (50%)
    sheet.cell(row=23, column=10).value = "=C23 * 1.0"  # Appliances - Month 7 (100%)
    
    # Landscaping and other - month 7-8
    sheet.cell(row=24, column=10).value = "=C24 * 0.3"  # Grading - Month 7 (30%)
    sheet.cell(row=24, column=11).value = "=C24 * 0.7"  # Grading - Month 8 (70%)
    sheet.cell(row=25, column=11).value = "=C25 * 1.0"  # Planting - Month 8 (100%)
    sheet.cell(row=26, column=11).value = "=C26 * 1.0"  # Cleaning - Month 8 (100%)
    
    # Dumpsters - spread across project
    sheet.cell(row=27, column=4).value = "=C27 * 0.1"  # Dumpsters - Month 1 (10%)
    sheet.cell(row=27, column=5).value = "=C27 * 0.1"  # Dumpsters - Month 2 (10%)
    sheet.cell(row=27, column=6).value = "=C27 * 0.1"  # Dumpsters - Month 3 (10%)
    sheet.cell(row=27, column=7).value = "=C27 * 0.15"  # Dumpsters - Month 4 (15%)
    sheet.cell(row=27, column=8).value = "=C27 * 0.15"  # Dumpsters - Month 5 (15%)
    sheet.cell(row=27, column=9).value = "=C27 * 0.15"  # Dumpsters - Month 6 (15%)
    sheet.cell(row=27, column=10).value = "=C27 * 0.15"  # Dumpsters - Month 7 (15%)
    sheet.cell(row=27, column=11).value = "=C27 * 0.1"  # Dumpsters - Month 8 (10%)
    
    # Contingency - distribute proportionally
    sheet.cell(row=28, column=4).value = "=SUM(D2:D27)/SUM(C2:C27)*C28"  # Contingency - Month 1
    sheet.cell(row=28, column=5).value = "=SUM(E2:E27)/SUM(C2:C27)*C28"  # Contingency - Month 2
    sheet.cell(row=28, column=6).value = "=SUM(F2:F27)/SUM(C2:C27)*C28"  # Contingency - Month 3
    sheet.cell(row=28, column=7).value = "=SUM(G2:G27)/SUM(C2:C27)*C28"  # Contingency - Month 4
    sheet.cell(row=28, column=8).value = "=SUM(H2:H27)/SUM(C2:C27)*C28"  # Contingency - Month 5
    sheet.cell(row=28, column=9).value = "=SUM(I2:I27)/SUM(C2:C27)*C28"  # Contingency - Month 6
    sheet.cell(row=28, column=10).value = "=SUM(J2:J27)/SUM(C2:C27)*C28"  # Contingency - Month 7
    sheet.cell(row=28, column=11).value = "=SUM(K2:K27)/SUM(C2:C27)*C28"  # Contingency - Month 8
    
    # Add monthly totals row
    row = rows  # This should be 29
    sheet.cell(row=row, column=1, value="MONTHLY TOTALS").font = Font(bold=True)
    sheet.cell(row=row, column=1).border = thin_border
    sheet.cell(row=row, column=2, value="").border = thin_border
    
    total_budget_cell = sheet.cell(row=row, column=3)
    total_budget_cell.value = f"=SUM(C2:C{row-1})"
    total_budget_cell.number_format = '$#,##0.00'
    total_budget_cell.font = Font(bold=True)
    total_budget_cell.border = thin_border
    
    # Sum each month's column
    for month in range(4, 12):  # Months 1-8
        month_total_cell = sheet.cell(row=row, column=month)
        month_total_cell.value = f"=SUM({get_column_letter(month)}2:{get_column_letter(month)}{row-1})"
        month_total_cell.number_format = '$#,##0.00'
        month_total_cell.font = Font(bold=True)
        month_total_cell.border = thin_border
    
    # Total forecast sum
    forecast_sum_cell = sheet.cell(row=row, column=12)
    forecast_sum_cell.value = f"=SUM(L2:L{row-1})"
    forecast_sum_cell.number_format = '$#,##0.00'
    forecast_sum_cell.font = Font(bold=True)
    forecast_sum_cell.border = thin_border
    
    # Variance sum
    variance_sum_cell = sheet.cell(row=row, column=13)
    variance_sum_cell.value = f"=SUM(M2:M{row-1})"
    variance_sum_cell.number_format = '$#,##0.00'
    variance_sum_cell.font = Font(bold=True)
    variance_sum_cell.border = thin_border
    
    # Add cumulative totals row
    row += 1
    sheet.cell(row=row, column=1, value="CUMULATIVE TOTALS").font = Font(bold=True)
    sheet.cell(row=row, column=1).border = thin_border
    sheet.cell(row=row, column=2, value="").border = thin_border
    sheet.cell(row=row, column=3, value="").border = thin_border
    
    # First month is just that month's total
    cum_month1_cell = sheet.cell(row=row, column=4)
    cum_month1_cell.value = f"=D{row-1}"
    cum_month1_cell.number_format = '$#,##0.00'
    cum_month1_cell.font = Font(bold=True)
    cum_month1_cell.border = thin_border
    
    # Subsequent months add previous cumulative total
    for month in range(5, 12):  # Months 2-8
        cum_month_cell = sheet.cell(row=row, column=month)
        prev_month_letter = get_column_letter(month-1)
        curr_month_letter = get_column_letter(month)
        cum_month_cell.value = f"={prev_month_letter}{row}+{curr_month_letter}{row-1}"
        cum_month_cell.number_format = '$#,##0.00'
        cum_month_cell.font = Font(bold=True)
        cum_month_cell.border = thin_border
    
    sheet.cell(row=row, column=12, value="").border = thin_border
    sheet.cell(row=row, column=13, value="").border = thin_border
    
    # Add percentage complete row
    row += 1
    sheet.cell(row=row, column=1, value="PERCENTAGE COMPLETE").font = Font(bold=True)
    sheet.cell(row=row, column=1).border = thin_border
    sheet.cell(row=row, column=2, value="").border = thin_border
    sheet.cell(row=row, column=3, value="").border = thin_border
    
    # Calculate percentage complete for each month
    for month in range(4, 12):  # Months 1-8
        pct_cell = sheet.cell(row=row, column=month)
        pct_cell.value = f"={get_column_letter(month)}{row-1}/C{row-2}"
        pct_cell.number_format = '0.00%'
        pct_cell.font = Font(bold=True)
        pct_cell.border = thin_border
    
    sheet.cell(row=row, column=12, value="").border = thin_border
    sheet.cell(row=row, column=13, value="").border = thin_border

def create_cost_tracking_sheet(wb, sheet, header_font, header_fill, alt_row_fill, thin_border):
    """Creates and formats the Cost Tracking sheet"""
    
    # Set column widths
    columns = [
        ('Category', 20), ('Subcategory', 25), ('Total Budget', 15),
        ('Committed Cost', 15), ('Invoiced to Date', 15), ('Paid to Date', 15),
        ('Remaining to Invoice', 18), ('Budget Variance', 15), ('Percent Complete', 15),
        ('Notes', 30)
    ]
    
    for i, (header, width) in enumerate(columns, start=1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width
        cell = sheet.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Link data from Budget Plan sheet
    budget_sheet = wb["Budget Plan"]
    rows = budget_sheet.max_row
    
    # For each row in Budget Plan, create a corresponding row in Cost Tracking
    for i in range(2, rows): # Skip header row and total row
        row = i
        
        # Copy category and subcategory
        sheet.cell(row=row, column=1, value=f"='Budget Plan'!A{row}").border = thin_border
        sheet.cell(row=row, column=2, value=f"='Budget Plan'!B{row}").border = thin_border
        
        # Link to budget total
        budget_cell = sheet.cell(row=row, column=3, value=f"='Budget Plan'!H{row}")
        budget_cell.number_format = '$#,##0.00'
        budget_cell.border = thin_border
        
        # Initialize cost tracking cells
        # Committed Cost - Will be filled in as contracts are awarded
        committed_cell = sheet.cell(row=row, column=4, value=0)
        committed_cell.number_format = '$#,##0.00'
        committed_cell.border = thin_border
        
        # Invoiced to Date - Will be filled in as invoices are received
        invoiced_cell = sheet.cell(row=row, column=5, value=0)
        invoiced_cell.number_format = '$#,##0.00'
        invoiced_cell.border = thin_border
        
        # Paid to Date - Will be filled in as payments are made
        paid_cell = sheet.cell(row=row, column=6, value=0)
        paid_cell.number_format = '$#,##0.00'
        paid_cell.border = thin_border
        
        # Formulas for calculated fields
        # Remaining to Invoice = Committed - Invoiced
        remaining_cell = sheet.cell(row=row, column=7)
        remaining_cell.value = f"=D{row}-E{row}"
        remaining_cell.number_format = '$#,##0.00'
        remaining_cell.border = thin_border
        
        # Budget Variance = Budget - Committed
        variance_cell = sheet.cell(row=row, column=8)
        variance_cell.value = f"=C{row}-D{row}"
        variance_cell.number_format = '$#,##0.00'
        variance_cell.border = thin_border
        
        # Percent Complete = Invoiced / Budget (if Budget > 0)
        complete_cell = sheet.cell(row=row, column=9)
        complete_cell.value = f"=IF(C{row}=0,0,E{row}/C{row})"
        complete_cell.number_format = '0.00%'
        complete_cell.border = thin_border
        
        # Notes field - empty
        notes_cell = sheet.cell(row=row, column=10, value="")
        notes_cell.border = thin_border
        
        # Apply alternating row formatting
        if row % 2 == 0:
            for col in range(1, 11):
                sheet.cell(row=row, column=col).fill = alt_row_fill
    
    # Add total row
    row = rows  # Should be 29 (after contingency)
    sheet.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    sheet.cell(row=row, column=1).border = thin_border
    sheet.cell(row=row, column=2, value="").border = thin_border
    
    # Sum the columns
    for col in range(3, 9):
        total_cell = sheet.cell(row=row, column=col)
        total_cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{row-1})"
        total_cell.number_format = '$#,##0.00'
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
    
    # Percent complete for project overall
    overall_complete_cell = sheet.cell(row=row, column=9)
    overall_complete_cell.value = f"=IF(C{row}=0,0,E{row}/C{row})"
    overall_complete_cell.number_format = '0.00%'
    overall_complete_cell.font = Font(bold=True)
    overall_complete_cell.border = thin_border
    
    sheet.cell(row=row, column=10, value="").border = thin_border

def create_dashboard_sheet(wb, sheet, header_font, header_fill, thin_border):
    """Creates and formats the Dashboard sheet"""
    
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 5
    sheet.column_dimensions['D'].width = 5
    sheet.column_dimensions['E'].width = 15
    
    # PROJECT SUMMARY
    sheet.merge_cells('A1:E1')
    summary_title = sheet.cell(row=1, column=1, value="PROJECT SUMMARY")
    summary_title.font = header_font
    summary_title.fill = header_fill
    summary_title.alignment = Alignment(horizontal='center')
    
    # Add borders to the merged cell
    for col in range(1, 6):
        sheet.cell(row=1, column=col).border = thin_border
    
    # Add budget summary
    sheet.cell(row=3, column=1, value="Total Budget").border = thin_border
    budget_cell = sheet.cell(row=3, column=2, value="='Budget Plan'!H29")
    budget_cell.number_format = '$#,##0.00'
    budget_cell.border = thin_border
    
    sheet.cell(row=4, column=1, value="Total Committed Cost").border = thin_border
    committed_cell = sheet.cell(row=4, column=2, value="='Cost Tracking'!D29")
    committed_cell.number_format = '$#,##0.00'
    committed_cell.border = thin_border
    
    sheet.cell(row=5, column=1, value="Total Invoiced to Date").border = thin_border
    invoiced_cell = sheet.cell(row=5, column=2, value="='Cost Tracking'!E29")
    invoiced_cell.number_format = '$#,##0.00'
    invoiced_cell.border = thin_border
    
    sheet.cell(row=6, column=1, value="Total Paid to Date").border = thin_border
    paid_cell = sheet.cell(row=6, column=2, value="='Cost Tracking'!F29")
    paid_cell.number_format = '$#,##0.00'
    paid_cell.border = thin_border
    
    sheet.cell(row=7, column=1, value="Budget Remaining").border = thin_border
    remaining_cell = sheet.cell(row=7, column=2, value="=B3-B4")
    remaining_cell.number_format = '$#,##0.00'
    remaining_cell.border = thin_border
    
    sheet.cell(row=8, column=1, value="Project Percent Complete").border = thin_border
    complete_cell = sheet.cell(row=8, column=2, value="='Cost Tracking'!I29")
    complete_cell.number_format = '0.00%'
    complete_cell.border = thin_border
    
    # Add percentage indicators
    for row in range(4, 8):
        sheet.cell(row=row, column=3).border = thin_border
        sheet.cell(row=row, column=4).border = thin_border
        
        pct_cell = sheet.cell(row=row, column=5)
        if row == 4:
            pct_cell.value = "=B4/B3"  # Committed / Budget
        elif row == 5:
            pct_cell.value = "=B5/B3"  # Invoiced / Budget
        elif row == 6:
            pct_cell.value = "=B6/B3"  # Paid / Budget
        elif row == 7:
            pct_cell.value = "=B7/B3"  # Remaining / Budget
        
        pct_cell.number_format = '0.00%'
        pct_cell.border = thin_border
    
    # Add borders to empty cells
    sheet.cell(row=3, column=3).border = thin_border
    sheet.cell(row=3, column=4).border = thin_border
    sheet.cell(row=3, column=5).border = thin_border
    sheet.cell(row=8, column=3).border = thin_border
    sheet.cell(row=8, column=4).border = thin_border
    sheet.cell(row=8, column=5).border = thin_border
    
    # SCHEDULE SUMMARY
    sheet.merge_cells('A10:E10')
    schedule_title = sheet.cell(row=10, column=1, value="SCHEDULE SUMMARY")
    schedule_title.font = header_font
    schedule_title.fill = header_fill
    schedule_title.alignment = Alignment(horizontal='center')
    
    # Add borders to the merged cell
    for col in range(1, 6):
        sheet.cell(row=10, column=col).border = thin_border
    
    # Add month headers
    months = ['Month 1', 'Month 2', 'Month 3', 'Month 4', 'Month 5', 'Month 6', 'Month 7', 'Month 8']
    for i, month in enumerate(months):
        row_num = 12 + i
        sheet.cell(row=row_num, column=1, value=month).border = thin_border
        
        month_cell = sheet.cell(row=row_num, column=2)
        month_col = chr(68 + i)  # 'D' for Month 1, 'E' for Month 2, etc.
        month_cell.value = f"='Budget Forecast'!{month_col}29"
        month_cell.number_format = '$#,##0.00'
        month_cell.border = thin_border
        
        sheet.cell(row=row_num, column=3).border = thin_border
        sheet.cell(row=row_num, column=4).border = thin_border
        
        pct_cell = sheet.cell(row=row_num, column=5)
        pct_cell.value = f"='Budget Forecast'!{month_col}31"
        pct_cell.number_format = '0.00%'
        pct_cell.border = thin_border
    
    # TOP EXPENSE CATEGORIES
    sheet.merge_cells('A21:E21')
    top_expenses_title = sheet.cell(row=21, column=1, value="TOP EXPENSE CATEGORIES")
    top_expenses_title.font = header_font
    top_expenses_title.fill = header_fill
    top_expenses_title.alignment = Alignment(horizontal='center')
    
    # Add borders to the merged cell
    for col in range(1, 6):
        sheet.cell(row=21, column=col).border = thin_border

def create_instructions_sheet(wb, sheet, header_font):
    """Creates and formats the Instructions sheet"""
    
    # Set column width
    sheet.column_dimensions['A'].width = 100
    
    # Add header
    header = sheet.cell(row=1, column=1, value="Construction Budget Workbook Instructions")
    header.font = header_font
    header.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    
    # Instructions text
    instructions = [
        "This workbook contains the following sheets to help manage your construction budget:",
        "",
        "1. Budget Plan: The main budget sheet where you enter all line items for your construction project.",
        "   - Enter your own categories, descriptions, quantities and unit prices",
        "   - The sheet will automatically calculate the total budget for each line item",
        "   - A contingency row is included and set to 10% of the total budget by default",
        "",
        "2. Budget Forecast: Shows the anticipated spending by month for each budget category.",
        "   - The sheet pulls data from the Budget Plan sheet",
        "   - Distribute expected costs across months by adjusting the percentage allocation for each line item",
        "   - Monthly and cumulative totals are calculated automatically",
        "   - Use this to plan cash flow needs throughout the project",
        "",
        "3. Cost Tracking: Track actual costs against budgeted amounts.",
        "   - Enter committed costs as contracts are awarded and purchase orders are issued",
        "   - Track invoiced and paid amounts",
        "   - Monitor budget variance (budgeted vs. committed)",
        "   - Track percent complete for each category",
        "   - Add notes to document changes or issues",
        "",
        "4. Dashboard: Provides a high-level overview of project budget status.",
        "   - Shows key budget metrics, including total budget, committed costs, and payments",
        "   - Displays monthly forecast and percentage complete",
        "   - Highlights top expense categories",
        "",
        "Tips for using this workbook:",
        "- Start by customizing the Budget Plan sheet with your specific line items",
        "- Update the Budget Forecast to distribute costs across the project timeline",
        "- Regularly update the Cost Tracking sheet with actual costs as they are committed and paid",
        "- Review the Dashboard for a quick project status overview",
        "- The sheet is linked with formulas, so changes in one area will update related sections automatically",
        "",
        "For best results:",
        "- Update the workbook at least weekly",
        "- Review with your project team regularly",
        "- Document changes in the Notes sections",
        "- Archive a copy of the workbook monthly to track changes over time"
    ]
    
    # Add instructions to the sheet
    for i, text in enumerate(instructions, start=2):
        sheet.cell(row=i, column=1, value=text)
        sheet.row_dimensions[i].height = 18

if __name__ == "__main__":
    create_construction_budget_workbook()
