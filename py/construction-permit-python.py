import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_construction_permit_log():
    """
    Creates a construction permitting log Excel file with sample data
    and statistics sheet.
    """
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Permit Log"
    
    # Define the header row for the permit log
    permits_header = [
        "Permit ID",
        "Project Name",
        "Project Address",
        "Permit Type",
        "Application Date",
        "Applicant Name",
        "Applicant Contact",
        "Reviewer Assigned",
        "Review Status",
        "Additional Info Requested",
        "Info Received Date",
        "Approval Date",
        "Permit Expiration",
        "Inspection Dates",
        "Inspection Results",
        "Final Approval Date",
        "Comments"
    ]
    
    # Create sample data for demonstration (5 rows of example data)
    sample_permit_data = [
        [
            "P-2025-001", 
            "Main Street Renovation", 
            "123 Main St, Anytown", 
            "Building Alteration", 
            "01/15/2025", 
            "John Smith", 
            "john.smith@email.com", 
            "Mary Johnson", 
            "Approved", 
            "No", 
            "", 
            "02/05/2025", 
            "08/05/2025", 
            "02/28/2025, 03/15/2025", 
            "Pass, Pass", 
            "03/20/2025", 
            "Project completed ahead of schedule"
        ],
        [
            "P-2025-002", 
            "Oak Plaza Development", 
            "456 Oak Ave, Anytown", 
            "New Construction", 
            "02/03/2025", 
            "Sarah Williams", 
            "swilliams@construction.com", 
            "Tom Lee", 
            "In Review", 
            "Yes", 
            "03/01/2025", 
            "", 
            "", 
            "", 
            "", 
            "", 
            "Structural calculations requested"
        ],
        [
            "P-2025-003", 
            "Elm Street Roofing", 
            "789 Elm St, Anytown", 
            "Roof Repair", 
            "02/10/2025", 
            "Robert Johnson", 
            "rjohnson@roofers.com", 
            "Linda Chen", 
            "Approved", 
            "No", 
            "", 
            "02/25/2025", 
            "08/25/2025", 
            "03/10/2025", 
            "Pass", 
            "03/12/2025", 
            ""
        ],
        [
            "P-2025-004", 
            "Maple Court Addition", 
            "321 Maple Ct, Anytown", 
            "Building Addition", 
            "02/18/2025", 
            "David Miller", 
            "dmiller@homes.net", 
            "Mary Johnson", 
            "Additional Info Required", 
            "Yes", 
            "", 
            "", 
            "", 
            "", 
            "", 
            "", 
            "Zoning variance needed"
        ],
        [
            "P-2025-005", 
            "Pine Street Demolition", 
            "567 Pine St, Anytown", 
            "Demolition", 
            "03/01/2025", 
            "Michael Brown", 
            "mbrown@demo.com", 
            "Tom Lee", 
            "Approved", 
            "No", 
            "", 
            "03/15/2025", 
            "09/15/2025", 
            "", 
            "", 
            "", 
            "Environmental clearance verified"
        ]
    ]
    
    # Add header row
    for col_idx, header in enumerate(permits_header, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    
    # Add sample data
    for row_idx, row_data in enumerate(sample_permit_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Set column widths for better readability
    column_widths = {
        1: 10,   # Permit ID
        2: 25,   # Project Name
        3: 30,   # Project Address
        4: 20,   # Permit Type
        5: 15,   # Application Date
        6: 20,   # Applicant Name
        7: 25,   # Applicant Contact
        8: 20,   # Reviewer Assigned
        9: 20,   # Review Status
        10: 15,  # Additional Info Requested
        11: 15,  # Info Received Date
        12: 15,  # Approval Date
        13: 15,  # Permit Expiration
        14: 20,  # Inspection Dates
        15: 20,  # Inspection Results
        16: 15,  # Final Approval Date
        17: 40,  # Comments
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Create a second sheet for permit statistics
    stats_ws = wb.create_sheet(title="Statistics")
    
    stats_header = ["Metric", "Count", "Percentage"]
    stats_data = [
        ["Total Permits", 5, "100%"],
        ["Approved Permits", 3, "60%"],
        ["Pending Review", 1, "20%"],
        ["Additional Info Required", 1, "20%"],
        ["Expired Permits", 0, "0%"]
    ]
    
    # Add header row to statistics sheet
    for col_idx, header in enumerate(stats_header, 1):
        cell = stats_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    
    # Add statistics data
    for row_idx, row_data in enumerate(stats_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            stats_ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Set column widths for statistics sheet
    stats_ws.column_dimensions['A'].width = 25
    stats_ws.column_dimensions['B'].width = 10
    stats_ws.column_dimensions['C'].width = 15
    
    # Save the workbook
    wb.save("Construction_Permitting_Log.xlsx")
    print("Construction Permitting Log created successfully!")

def create_permit_application_form():
    """
    Creates a construction permit application form Excel file.
    """
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Permit Application"
    
    # Define the sections and fields for the application form
    application_form = [
        # Title and Instructions
        ["CONSTRUCTION PERMIT APPLICATION FORM"],
        ["Please complete all sections of this form. Fields marked with an asterisk (*) are required."],
        [""],
        
        # Section 1: Applicant Information
        ["SECTION 1: APPLICANT INFORMATION"],
        ["*Applicant Name:", ""],
        ["Company Name:", ""],
        ["*Mailing Address:", ""],
        ["*City:", "", "*State:", "", "*ZIP Code:", ""],
        ["*Phone Number:", "", "Email:", ""],
        ["Contractor License #:", "", "License Type:", ""],
        [""],
        
        # Section 2: Project Information
        ["SECTION 2: PROJECT INFORMATION"],
        ["*Project Address:", ""],
        ["*City:", "", "*State:", "", "*ZIP Code:", ""],
        ["*APN/Parcel Number:", ""],
        ["Zoning District:", ""],
        ["*Estimated Project Cost: $", ""],
        ["*Project Description:", ""],
        ["", ""],  # Extra space for description
        ["", ""],  # Extra space for description
        [""],
        
        # Section 3: Project Type
        ["SECTION 3: PROJECT TYPE (Check all that apply)"],
        ["[ ] New Construction", "", "[ ] Renovation/Remodel", ""],
        ["[ ] Addition", "", "[ ] Demolition", ""],
        ["[ ] Plumbing", "", "[ ] Electrical", ""],
        ["[ ] Mechanical", "", "[ ] Roofing", ""],
        ["[ ] Grading/Excavation", "", "[ ] Other (specify):", ""],
        [""],
        
        # Section 4: Building Information
        ["SECTION 4: BUILDING INFORMATION"],
        ["*Occupancy Type:", ""],
        ["*Construction Type:", ""],
        ["*Number of Stories:", "", "*Building Height:", ""],
        ["*Total Square Footage:", ""],
        ["Existing Square Footage:", "", "New Square Footage:", ""],
        ["Number of Bedrooms:", "", "Number of Bathrooms:", ""],
        ["Fire Sprinklers: [ ] Yes  [ ] No", ""],
        [""],
        
        # Section 5: Required Attachments
        ["SECTION 5: REQUIRED ATTACHMENTS"],
        ["Check all items included with this application:"],
        ["[ ] Site Plan", "", "[ ] Floor Plan", ""],
        ["[ ] Elevation Drawings", "", "[ ] Structural Calculations", ""],
        ["[ ] Environmental Documents", "", "[ ] Title 24 Energy Calculations", ""],
        ["[ ] Soils Report", "", "[ ] Other:", ""],
        [""],
        
        # Section 6: Declarations and Signatures
        ["SECTION 6: DECLARATIONS AND SIGNATURES"],
        ["I hereby certify that I have read and examined this application and know the same to be true and correct. All provisions of laws and ordinances governing this type of work will be complied with whether specified herein or not. The granting of a permit does not presume to give authority to violate or cancel the provisions of any federal, state, or local law regulating construction or the performance of construction."],
        [""],
        ["*Owner/Authorized Agent Signature:", "", "*Date:", ""],
        ["*Print Name:", "", "*Title:", ""],
        [""],
        
        # For Official Use Only
        ["FOR OFFICIAL USE ONLY"],
        ["Permit #:", "", "Date Received:", ""],
        ["Received By:", "", "Fee Amount: $", ""],
        ["Review Required: [ ] Building  [ ] Planning  [ ] Engineering  [ ] Fire  [ ] Health"],
        ["Comments:", ""],
        ["", ""],  # Extra space for comments
        ["", ""]   # Extra space for comments
    ]
    
    # Set initial row and populate form content
    current_row = 1
    merge_regions = []
    
    # Insert application form content row by row
    for row_data in application_form:
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
            
            # Style for title
            if current_row == 1 and col_idx == 1:
                cell.font = Font(bold=True, size=16)
                cell.alignment = Alignment(horizontal='center')
                merge_regions.append((current_row, 1, current_row, 4))
            
            # Style for section headers
            elif cell_value.startswith("SECTION") or cell_value == "FOR OFFICIAL USE ONLY":
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                merge_regions.append((current_row, 1, current_row, 4))
        
        current_row += 1
    
    # Apply the merge regions
    for region in merge_regions:
        start_row, start_col, end_row, end_col = region
        ws.merge_cells(
            start_row=start_row, 
            start_column=start_col, 
            end_row=end_row, 
            end_column=end_col
        )
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    
    # Set row heights for title and section headers
    for row_idx in range(1, current_row):
        if row_idx == 1:  # Title row
            ws.row_dimensions[row_idx].height = 30
        elif row_idx == 4 or row_idx == 12 or row_idx == 22 or row_idx == 29 or row_idx == 38 or row_idx == 46 or row_idx == 53:
            # Section headers
            ws.row_dimensions[row_idx].height = 22
        else:
            # Regular rows
            ws.row_dimensions[row_idx].height = 18
    
    # Save the workbook
    wb.save("Construction_Permit_Application.xlsx")
    print("Construction Permit Application Form created successfully!")

# Generate both files when the script is run
if __name__ == "__main__":
    create_construction_permit_log()
    create_permit_application_form()
    print("Both Excel files have been created successfully!")
