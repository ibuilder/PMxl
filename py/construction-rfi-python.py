#!/usr/bin/env python3
# Construction RFI Log and Form Generator
# Requires: pip install openpyxl

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()

# ===== SHEET 1: RFI LOG =====
log_sheet = wb.active
log_sheet.title = "RFI Log"
log_sheet.sheet_properties.tabColor = "4472C4"

# Define column headers
headers = [
    'RFI #', 'Date Submitted', 'Title', 'Description', 'Requested By', 
    'Company', 'Discipline', 'Priority', 'Status', 'Assigned To', 
    'Response Date', 'Days Open', 'File Attachments'
]

# Set column widths
column_widths = [8, 15, 30, 50, 20, 20, 15, 10, 12, 20, 15, 10, 30]
for i, width in enumerate(column_widths, 1):
    log_sheet.column_dimensions[get_column_letter(i)].width = width

# Add header row
for col_num, header in enumerate(headers, 1):
    cell = log_sheet.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add sample data
sample_data = [
    'RFI-001',
    datetime.now(),
    'Foundation Depth Clarification',
    'Need confirmation on the foundation depth for column locations B3 through B8',
    'John Smith',
    'ABC Construction',
    'Structural',
    'High',
    'Open',
    'Jane Architect',
    '',
    '=IF(K2="",TODAY()-B2,K2-B2)',
    'Foundation-plan-rev2.pdf'
]

for col_num, value in enumerate(sample_data, 1):
    log_sheet.cell(row=2, column=col_num).value = value

# Create data validations
# Status validation
status_validation = DataValidation(
    type="list",
    formula1='"Open,In Review,Responded,Closed"',
    allow_blank=False
)
log_sheet.add_data_validation(status_validation)
status_validation.add(f'I2:I1000')

# Priority validation
priority_validation = DataValidation(
    type="list",
    formula1='"Low,Medium,High,Critical"',
    allow_blank=False
)
log_sheet.add_data_validation(priority_validation)
priority_validation.add(f'H2:H1000')

# Discipline validation
discipline_validation = DataValidation(
    type="list",
    formula1='"Architectural,Structural,Mechanical,Electrical,Plumbing,Civil,Other"',
    allow_blank=False
)
log_sheet.add_data_validation(discipline_validation)
discipline_validation.add(f'G2:G1000')

# Apply alternating row colors (stripes) for readability
for row in range(2, 1000):
    if row % 2 == 0:
        for col in range(1, len(headers) + 1):
            log_sheet.cell(row=row, column=col).fill = PatternFill(
                start_color="E9EFF7", end_color="E9EFF7", fill_type="solid"
            )

# ===== SHEET 2: RFI FORM =====
form_sheet = wb.create_sheet(title="RFI Form")
form_sheet.sheet_properties.tabColor = "70AD47"

# Set column widths
form_sheet.column_dimensions['A'].width = 20
form_sheet.column_dimensions['B'].width = 50
form_sheet.column_dimensions['C'].width = 20
form_sheet.column_dimensions['D'].width = 20

# Define common styles
title_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
header_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
logo_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
bold_font = Font(bold=True)
title_font = Font(bold=True, size=16)
header_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add company logo placeholder
form_sheet.merge_cells('A1:B3')
form_sheet['A1'] = 'COMPANY LOGO'
form_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
form_sheet['A1'].font = title_font
form_sheet['A1'].fill = logo_fill

# RFI Form Title
form_sheet.merge_cells('C1:D3')
form_sheet['C1'] = 'REQUEST FOR INFORMATION'
form_sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
form_sheet['C1'].font = title_font
form_sheet['C1'].fill = title_fill

# Add horizontal separator
form_sheet.merge_cells('A4:D4')
form_sheet.row_dimensions[4].height = 5
form_sheet['A4'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

# RFI Details Section
form_sheet.merge_cells('A5:D5')
form_sheet['A5'] = 'RFI DETAILS'
form_sheet['A5'].font = header_font
form_sheet['A5'].fill = header_fill

# RFI Metadata
form_sheet['A6'] = 'RFI Number:'
form_sheet['A6'].font = bold_font
form_sheet['B6'] = 'RFI-XXX'

form_sheet['C6'] = 'Date Submitted:'
form_sheet['C6'].font = bold_font
form_sheet['D6'] = datetime.now()
form_sheet['D6'].number_format = 'mm/dd/yyyy'

form_sheet['A7'] = 'Project:'
form_sheet['A7'].font = bold_font
form_sheet['B7'] = ''

form_sheet['C7'] = 'Priority:'
form_sheet['C7'].font = bold_font

# Priority dropdown
priority_validation = DataValidation(
    type="list",
    formula1='"Low,Medium,High,Critical"',
    allow_blank=False
)
form_sheet.add_data_validation(priority_validation)
priority_validation.add('D7')

form_sheet['A8'] = 'Requested By:'
form_sheet['A8'].font = bold_font
form_sheet['B8'] = ''

form_sheet['C8'] = 'Company:'
form_sheet['C8'].font = bold_font
form_sheet['D8'] = ''

form_sheet['A9'] = 'Discipline:'
form_sheet['A9'].font = bold_font

# Discipline dropdown
discipline_validation = DataValidation(
    type="list",
    formula1='"Architectural,Structural,Mechanical,Electrical,Plumbing,Civil,Other"',
    allow_blank=False
)
form_sheet.add_data_validation(discipline_validation)
discipline_validation.add('B9')

form_sheet['C9'] = 'Required Date:'
form_sheet['C9'].font = bold_font
form_sheet['D9'].number_format = 'mm/dd/yyyy'

# Request Section
form_sheet.merge_cells('A10:D10')
form_sheet['A10'] = 'REQUEST'
form_sheet['A10'].font = header_font
form_sheet['A10'].fill = header_fill

form_sheet['A11'] = 'Request Title:'
form_sheet['A11'].font = bold_font
form_sheet.merge_cells('B11:D11')

form_sheet['A12'] = 'Specification Reference:'
form_sheet['A12'].font = bold_font
form_sheet.merge_cells('B12:D12')

form_sheet['A13'] = 'Drawing Reference:'
form_sheet['A13'].font = bold_font
form_sheet.merge_cells('B13:D13')

form_sheet['A14'] = 'Description of Request:'
form_sheet['A14'].font = bold_font
form_sheet.merge_cells('A15:D19')
form_sheet['A15'].alignment = Alignment(wrapText=True, vertical='top')
form_sheet['A15'].border = thin_border

form_sheet['A20'] = 'Attachments:'
form_sheet['A20'].font = bold_font
form_sheet.merge_cells('B20:D20')

# Response Section
form_sheet.merge_cells('A21:D21')
form_sheet['A21'] = 'RESPONSE (To be completed by Design Team)'
form_sheet['A21'].font = header_font
form_sheet['A21'].fill = header_fill

form_sheet['A22'] = 'Responded By:'
form_sheet['A22'].font = bold_font
form_sheet['B22'] = ''

form_sheet['C22'] = 'Response Date:'
form_sheet['C22'].font = bold_font
form_sheet['D22'].number_format = 'mm/dd/yyyy'

form_sheet.merge_cells('A23:D23')
form_sheet['A23'] = 'Response:'
form_sheet['A23'].font = bold_font

form_sheet.merge_cells('A24:D28')
form_sheet['A24'].alignment = Alignment(wrapText=True, vertical='top')
form_sheet['A24'].border = thin_border

form_sheet['A29'] = 'Response Attachments:'
form_sheet['A29'].font = bold_font
form_sheet.merge_cells('B29:D29')

# Additional Comments Section
form_sheet.merge_cells('A30:D30')
form_sheet['A30'] = 'ADDITIONAL COMMENTS'
form_sheet['A30'].font = header_font
form_sheet['A30'].fill = header_fill

form_sheet.merge_cells('A31:D34')
form_sheet['A31'].alignment = Alignment(wrapText=True, vertical='top')
form_sheet['A31'].border = thin_border

# Add instructions
form_sheet.merge_cells('A36:D37')
form_sheet['A36'] = 'INSTRUCTIONS: Complete all fields in the RFI DETAILS and REQUEST sections. Submit to the architect/engineer for response. Use the log sheet to track all RFIs.'
form_sheet['A36'].alignment = Alignment(wrapText=True)
form_sheet['A36'].font = Font(italic=True)

# Save the workbook
wb.save('Construction_RFI_Template.xlsx')

print("Construction RFI Template has been created successfully.")
