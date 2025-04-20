import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import datetime
import os

def create_construction_schedule():
    """
    Creates a complete Excel workbook for a construction project schedule
    with multiple sheets for task list, Gantt chart, milestones,
    constraints, and logistics phasing.
    """
    # Create a Pandas Excel writer
    filename = 'Construction_Project_Schedule.xlsx'
    
    # Initialize the workbook
    wb = Workbook()
    
    # Define border style for headers
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define header style
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ------ TASK LIST SHEET ------
    # Create task list data
    task_data = {
        'ID': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17],
        'Task Name': [
            'Project Initiation', 'Site Preparation', 'Foundation Work', 'Structural Framing', 
            'Roof Installation', 'Exterior Walls', 'Plumbing Rough-In', 'Electrical Rough-In', 
            'HVAC Installation', 'Insulation', 'Drywall', 'Interior Finishes', 
            'Final Plumbing', 'Final Electrical', 'Site Cleanup', 'Final Inspection', 'Project Handover'
        ],
        'Description': [
            'Initial project setup and paperwork',
            'Clear the site and prepare for foundation work',
            'Excavation and foundation construction',
            'Construct main building frame',
            'Install roofing system',
            'Construct exterior walls and cladding',
            'Initial plumbing installation',
            'Initial electrical systems installation',
            'Install heating and cooling systems',
            'Install building insulation',
            'Install interior wall panels',
            'Paint, trim, flooring and fixtures',
            'Complete plumbing fixtures and connections',
            'Complete electrical fixtures and connections',
            'Final site cleanup and preparation',
            'Regulatory inspections and approvals',
            'Complete documentation and client handover'
        ],
        'Responsible': [
            'Project Manager', 'Site Supervisor', 'Civil Engineer', 'Structural Engineer',
            'Roofing Contractor', 'Construction Team', 'Plumbing Contractor', 'Electrical Contractor',
            'HVAC Contractor', 'Insulation Contractor', 'Drywall Contractor', 'Interior Contractor',
            'Plumbing Contractor', 'Electrical Contractor', 'Site Supervisor', 'Project Manager', 'Project Manager'
        ],
        'Duration (days)': [10, 15, 20, 30, 15, 20, 15, 15, 20, 10, 15, 25, 10, 10, 5, 5, 3],
        'Start Date': [
            '2025-05-01', '2025-05-15', '2025-06-05', '2025-07-03', '2025-08-14', '2025-08-14',
            '2025-09-11', '2025-09-11', '2025-09-11', '2025-10-09', '2025-10-23', '2025-11-13',
            '2025-12-18', '2025-12-18', '2026-01-01', '2026-01-08', '2026-01-15'
        ],
        'End Date': [
            '2025-05-14', '2025-06-04', '2025-07-02', '2025-08-13', '2025-09-03', '2025-09-10',
            '2025-10-01', '2025-10-01', '2025-10-08', '2025-10-22', '2025-11-12', '2025-12-17',
            '2025-12-31', '2025-12-31', '2026-01-07', '2026-01-14', '2026-01-19'
        ],
        'Dependencies': [
            '', '1', '2', '3', '4', '4', '4', '4', '4', '6,7,8,9', '10', '11', '12', '12', '13,14', '15', '16'
        ],
        'Status': ['Not Started'] * 17,
        'Priority': [
            'High', 'High', 'High', 'High', 'Medium', 'Medium', 'Medium', 'Medium',
            'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Medium', 'Low', 'High', 'High'
        ],
        'Notes': [''] * 17
    }
    
    # Create DataFrame
    task_df = pd.DataFrame(task_data)
    
    # Convert dates to datetime
    task_df['Start Date'] = pd.to_datetime(task_df['Start Date'])
    task_df['End Date'] = pd.to_datetime(task_df['End Date'])
    
    # Delete the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create and format the Task List sheet
    ws_tasks = wb.create_sheet('Task List')
    
    # Write headers
    headers = list(task_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws_tasks.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    
    # Write data
    for row_num, row_data in enumerate(task_df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_tasks.cell(row=row_num, column=col_num)
            if isinstance(cell_value, pd.Timestamp):
                cell.value = cell_value.strftime('%Y-%m-%d')
            else:
                cell.value = cell_value
            cell.alignment = Alignment(vertical='center')
    
    # Set column widths
    column_widths = [5, 20, 35, 20, 15, 12, 12, 15, 15, 10, 25]
    for i, width in enumerate(column_widths, 1):
        ws_tasks.column_dimensions[get_column_letter(i)].width = width
    
    # ------ GANTT CHART SHEET ------
    ws_gantt = wb.create_sheet('Gantt Chart')
    
    # Gantt chart headers
    gantt_headers = ['ID', 'Task Name', 'Start Date', 'End Date', 'Duration (days)']
    for col_num, header in enumerate(gantt_headers, 1):
        cell = ws_gantt.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    
    # Calendar headers - Add days for 9 months (36 weeks)
    # Start with project start date
    start_date = pd.to_datetime(task_df['Start Date'].min())
    
    # Generate calendar dates (weeks) for the duration of the project
    calendar_start = 6  # Column F starts the calendar
    current_date = start_date
    end_date = pd.to_datetime(task_df['End Date'].max())
    week_num = 1
    
    while current_date <= end_date:
        col = calendar_start + week_num - 1
        cell = ws_gantt.cell(row=1, column=col)
        cell.value = f"Week {week_num}\n{current_date.strftime('%m/%d')}"
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # Move to next week
        current_date += pd.Timedelta(days=7)
        week_num += 1
    
    # Write Gantt data
    for row_num, (_, row) in enumerate(task_df.iterrows(), 2):
        # Basic task data
        ws_gantt.cell(row=row_num, column=1).value = row['ID']
        ws_gantt.cell(row=row_num, column=2).value = row['Task Name']
        ws_gantt.cell(row=row_num, column=3).value = row['Start Date'].strftime('%Y-%m-%d')
        ws_gantt.cell(row=row_num, column=4).value = row['End Date'].strftime('%Y-%m-%d')
        ws_gantt.cell(row=row_num, column=5).value = row['Duration (days)']
        
        # Calculate bar position
        task_start = pd.to_datetime(row['Start Date'])
        task_end = pd.to_datetime(row['End Date'])
        
        start_week = ((task_start - start_date).days // 7) + 1
        end_week = ((task_end - start_date).days // 7) + 1
        
        # Fill cells for the duration
        for week in range(start_week, end_week + 1):
            col = calendar_start + week - 1
            cell = ws_gantt.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
    
    # Set column widths
    ws_gantt.column_dimensions[get_column_letter(1)].width = 5
    ws_gantt.column_dimensions[get_column_letter(2)].width = 25
    ws_gantt.column_dimensions[get_column_letter(3)].width = 12
    ws_gantt.column_dimensions[get_column_letter(4)].width = 12
    ws_gantt.column_dimensions[get_column_letter(5)].width = 12
    
    # Set weekly column widths
    for i in range(calendar_start, calendar_start + week_num):
        ws_gantt.column_dimensions[get_column_letter(i)].width = 8
    
    # ------ MILESTONES SHEET ------
    milestone_data = {
        'ID': ['M1', 'M2', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8'],
        'Milestone Name': [
            'Project Start', 'Foundation Complete', 'Structure Complete', 'Building Enclosed',
            'MEP Rough-in Complete', 'Interior Finishes Complete', 'Final Inspections', 'Project Completion'
        ],
        'Target Date': [
            '2025-05-01', '2025-07-02', '2025-08-13', '2025-09-10',
            '2025-10-08', '2025-12-17', '2026-01-14', '2026-01-19'
        ],
        'Responsible': [
            'Project Manager', 'Civil Engineer', 'Structural Engineer', 'Construction Team',
            'MEP Coordinator', 'Interior Contractor', 'Project Manager', 'Project Manager'
        ],
        'Associated Tasks': ['1', '3', '4', '5,6', '7,8,9', '12', '16', '17'],
        'Status': ['Not Started'] * 8,
        'Notes': [''] * 8
    }
    
    milestone_df = pd.DataFrame(milestone_data)
    
    # Create milestone sheet
    ws_milestones = wb.create_sheet('Milestones')
    
    # Write headers
    headers = list(milestone_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws_milestones.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    
    # Write data
    for row_num, row_data in enumerate(milestone_df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws_milestones.cell(row=row_num, column=col_num).value = cell_value
    
    # Set column widths
    milestone_widths = [5, 25, 12, 20, 15, 15, 25]
    for i, width in enumerate(milestone_widths, 1):
        ws_milestones.column_dimensions[get_column_letter(i)].width = width
    
    # ------ CONSTRAINTS SHEET ------
    constraint_data = {
        'ID': ['C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7'],
        'Constraint Type': [
            'Weather', 'Regulatory', 'Resource', 'Budget', 'Site', 'Technical', 'Environmental'
        ],
        'Description': [
            'Winter conditions may delay exterior work',
            'Building permit approval process',
            'Limited skilled labor availability',
            'Material cost fluctuations',
            'Limited site access for deliveries',
            'Complex foundation requirements',
            'Noise restrictions during certain hours'
        ],
        'Impact': [
            'Schedule delay',
            'Cannot start construction without permits',
            'Potential delays in specialized work',
            'Budget overruns',
            'Logistics complications',
            'Additional engineering required',
            'Limited working hours'
        ],
        'Affected Tasks': ['5,6', '2,3', '7,8,9', 'All', 'All', '3', 'All'],
        'Mitigation Plan': [
            'Schedule exterior work in warmer months, prepare contingency plan',
            'Submit applications early, follow up regularly',
            'Pre-book contractors, consider alternative sourcing',
            'Secure price commitments early, include contingency budget',
            'Create detailed delivery schedule, coordinate with neighbors',
            'Early engagement with geotechnical experts',
            'Schedule noisy work during permitted hours, notify community'
        ],
        'Status': ['Active'] * 7,
        'Responsible': [
            'Project Manager', 'Permit Coordinator', 'Resource Manager', 'Financial Manager',
            'Site Supervisor', 'Civil Engineer', 'Site Supervisor'
        ]
    }
    
    constraint_df = pd.DataFrame(constraint_data)
    
    # Create constraints sheet
    ws_constraints = wb.create_sheet('Constraints')
    
    # Write headers
    headers = list(constraint_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws_constraints.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    
    # Write data
    for row_num, row_data in enumerate(constraint_df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_constraints.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Set column widths
    constraint_widths = [5, 15, 35, 25, 15, 35, 10, 20]
    for i, width in enumerate(constraint_widths, 1):
        ws_constraints.column_dimensions[get_column_letter(i)].width = width
    
    # ------ LOGISTIC PHASING SHEET ------
    phasing_data = {
        'Phase': ['Phase 1', 'Phase 2', 'Phase 3', 'Phase 4', 'Phase 5'],
        'Start Date': ['2025-05-01', '2025-07-03', '2025-09-11', '2025-10-23', '2026-01-01'],
        'End Date': ['2025-07-02', '2025-09-10', '2025-10-22', '2025-12-31', '2026-01-19'],
        'Description': [
            'Site Preparation & Foundation',
            'Structure & Envelope',
            'MEP Rough-In & Insulation',
            'Interior Finishes & MEP Completion',
            'Completion & Handover'
        ],
        'Tasks Involved': ['1,2,3', '4,5,6', '7,8,9,10', '11,12,13,14', '15,16,17'],
        'Resources Required': [
            'Excavators, Concrete trucks, Laborers',
            'Crane, Delivery trucks, Framing team',
            'Specialized contractors, Material deliveries',
            'Finish contractors, Fixtures deliveries',
            'Cleaning crews, Inspection teams'
        ],
        'Site Access Points': [
            'Main entrance',
            'Main and east entrances',
            'All entrances',
            'Main entrance only',
            'Restricted access'
        ],
        'Storage Areas': [
            'North corner of site',
            'East side of building',
            'Interior of building',
            'Interior secured rooms',
            'Minimal on-site storage'
        ],
        'Special Requirements': [
            'Temporary fencing, Silt control',
            'Staging area for materials, Tower crane setup',
            'Secure storage for equipment, Temporary power',
            'Climate control active, Dust control measures',
            'Furniture deliveries, Systems testing'
        ]
    }
    
    phasing_df = pd.DataFrame(phasing_data)
    
    # Create phasing sheet
    ws_phasing = wb.create_sheet('Logistic Phasing')
    
    # Write headers
    headers = list(phasing_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws_phasing.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = thin_border
    
    # Write data
    for row_num, row_data in enumerate(phasing_df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_phasing.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Set column widths
    phasing_widths = [10, 12, 12, 25, 15, 30, 20, 20, 35]
    for i, width in enumerate(phasing_widths, 1):
        ws_phasing.column_dimensions[get_column_letter(i)].width = width
    
    # Set tab colors
    ws_tasks.sheet_properties.tabColor = "4472C4"  # Blue
    ws_gantt.sheet_properties.tabColor = "70AD47"  # Green
    ws_milestones.sheet_properties.tabColor = "ED7D31"  # Orange
    ws_constraints.sheet_properties.tabColor = "A5A5A5"  # Gray
    ws_phasing.sheet_properties.tabColor = "5B9BD5"  # Light Blue
    
    # Save the workbook
    wb.save(filename)
    print(f"Construction schedule created successfully: {filename}")
    return filename

if __name__ == "__main__":
    create_construction_schedule()
