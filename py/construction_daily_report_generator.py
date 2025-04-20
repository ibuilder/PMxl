import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import os

def create_construction_daily_report():
    # Create a workbook and get active worksheet
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create all required sheets
    dashboard = wb.create_sheet("Dashboard")
    daily_report = wb.create_sheet("Daily Report Form")
    reports_log = wb.create_sheet("Reports Log")
    weather_data = wb.create_sheet("Weather Data")
    manpower = wb.create_sheet("Manpower Tracking")
    equipment = wb.create_sheet("Equipment Log")
    materials = wb.create_sheet("Material Deliveries")
    photos = wb.create_sheet("Photos Index")
    schedule = wb.create_sheet("Schedule Tracking")
    
    # Set Dashboard as the first sheet
    wb.active = dashboard
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2A5885', end_color='2A5885', fill_type='solid')
    subheader_font = Font(name='Arial', size=11, bold=True)
    subheader_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Configure Dashboard
    configure_dashboard(dashboard, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Daily Report Form
    configure_daily_report_form(daily_report, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Reports Log
    configure_reports_log(reports_log, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Weather Data sheet
    configure_weather_data(weather_data, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Manpower Tracking sheet
    configure_manpower_tracking(manpower, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Equipment Log sheet
    configure_equipment_log(equipment, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Material Deliveries sheet
    configure_material_deliveries(materials, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Photos Index sheet
    configure_photos_index(photos, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Configure Schedule Tracking sheet
    configure_schedule_tracking(schedule, header_font, header_fill, subheader_font, subheader_fill, border)
    
    # Set column widths for all sheets
    for sheet in wb.worksheets:
        for col in range(1, 20):
            column = get_column_letter(col)
            sheet.column_dimensions[column].width = 15
    
    # Save the workbook
    wb.save("Construction_Daily_Report_Tracker.xlsx")
    print("Excel workbook 'Construction_Daily_Report_Tracker.xlsx' has been created successfully.")
    
    return "Construction_Daily_Report_Tracker.xlsx"

def configure_dashboard(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Set title
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "CONSTRUCTION PROJECT DAILY REPORTING DASHBOARD"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Project Information Section
    sheet.merge_cells('A3:F3')
    sheet['A3'] = "PROJECT INFORMATION"
    sheet['A3'].font = subheader_font
    sheet['A3'].fill = subheader_fill
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Project details
    sheet['A4'] = "Project Name:"
    sheet['B4'] = "[Enter Project Name]"
    sheet['D4'] = "Start Date:"
    sheet['E4'] = datetime.now().strftime("%m/%d/%Y")
    
    sheet['A5'] = "Project Number:"
    sheet['B5'] = "[Enter Project Number]"
    sheet['D5'] = "End Date:"
    sheet['E5'] = (datetime.now() + timedelta(days=365)).strftime("%m/%d/%Y")
    
    sheet['A6'] = "Location:"
    sheet['B6'] = "[Enter Location]"
    sheet['D6'] = "Current Date:"
    sheet['E6'] = "=TODAY()"
    sheet['E6'].number_format = "mm/dd/yyyy"
    
    sheet['A7'] = "Project Manager:"
    sheet['B7'] = "[Enter PM Name]"
    sheet['D7'] = "Days Elapsed:"
    sheet['E7'] = "=NETWORKDAYS(E4,E6)"
    
    # Quick Navigation Section
    sheet.merge_cells('A9:F9')
    sheet['A9'] = "QUICK NAVIGATION"
    sheet['A9'].font = subheader_font
    sheet['A9'].fill = subheader_fill
    sheet['A9'].alignment = Alignment(horizontal='center')
    
    # Navigation buttons (using hyperlinks)
    buttons = [
        ("A10", "Daily Report Form"), 
        ("C10", "Reports Log"),
        ("E10", "Weather Data"),
        ("A11", "Manpower Tracking"),
        ("C11", "Equipment Log"),
        ("E11", "Material Deliveries"),
        ("A12", "Photos Index"),
        ("C12", "Schedule Tracking")
    ]
    
    for cell, sheet_name in buttons:
        sheet[cell] = sheet_name
        sheet[cell].hyperlink = f"#{sheet_name}!A1"
        sheet[cell].font = Font(color="0563C1", underline="single")
        sheet[cell].alignment = Alignment(horizontal='center')
    
    # Project Summary Section
    sheet.merge_cells('A14:F14')
    sheet['A14'] = "PROJECT SUMMARY"
    sheet['A14'].font = subheader_font
    sheet['A14'].fill = subheader_fill
    sheet['A14'].alignment = Alignment(horizontal='center')
    
    # Summary metrics
    sheet['A15'] = "Total Man-hours:"
    sheet['B15'] = "=SUM('Manpower Tracking'!F:F)"
    
    sheet['D15'] = "Weather Delays:"
    sheet['E15'] = "=COUNTIFS('Weather Data'!E:E,\"Yes\")"
    
    sheet['A16'] = "Safety Incidents:"
    sheet['B16'] = "0"
    
    sheet['D16'] = "Material Deliveries:"
    sheet['E16'] = "=COUNTA('Material Deliveries'!A:A)-1"
    
    # Recent Reports
    sheet.merge_cells('A18:F18')
    sheet['A18'] = "RECENT DAILY REPORTS"
    sheet['A18'].font = subheader_font
    sheet['A18'].fill = subheader_fill
    sheet['A18'].alignment = Alignment(horizontal='center')
    
    # Headers for recent reports table
    headers = ["Date", "Report #", "Weather", "Manpower Count", "Delays"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=19, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
    
    # Reference formulas for recent reports (last 5)
    for row in range(20, 25):
        sheet.cell(row=row, column=1).value = f"=IFERROR(INDEX('Reports Log'!A:A,COUNTA('Reports Log'!A:A)-{25-row}),\"\")"
        sheet.cell(row=row, column=1).number_format = "mm/dd/yyyy"
        sheet.cell(row=row, column=2).value = f"=IFERROR(INDEX('Reports Log'!B:B,COUNTA('Reports Log'!A:A)-{25-row}),\"\")"
        sheet.cell(row=row, column=3).value = f"=IFERROR(INDEX('Reports Log'!D:D,COUNTA('Reports Log'!A:A)-{25-row}),\"\")"
        sheet.cell(row=row, column=4).value = f"=IFERROR(INDEX('Reports Log'!E:E,COUNTA('Reports Log'!A:A)-{25-row}),\"\")"
        sheet.cell(row=row, column=5).value = f"=IFERROR(INDEX('Reports Log'!G:G,COUNTA('Reports Log'!A:A)-{25-row}),\"\")"

def configure_daily_report_form(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:H1')
    sheet['A1'] = "DAILY CONSTRUCTION REPORT"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Basic information
    sheet['A3'] = "Date:"
    sheet['B3'] = datetime.now().strftime("%m/%d/%Y")
    sheet['B3'].number_format = "mm/dd/yyyy"
    
    sheet['C3'] = "Day:"
    sheet['D3'] = "=TEXT(B3,\"dddd\")"
    
    sheet['E3'] = "Report #:"
    sheet['F3'] = "=IF(B3=\"\",\"\",CONCATENATE(TEXT(B3,\"yyyymmdd\"),\"-\",ROW()))"
    
    sheet['G3'] = "Superintendent:"
    sheet['H3'] = "[Name]"
    
    # Weather section
    sheet.merge_cells('A5:H5')
    sheet['A5'] = "WEATHER CONDITIONS"
    sheet['A5'].font = subheader_font
    sheet['A5'].fill = subheader_fill
    sheet['A5'].alignment = Alignment(horizontal='center')
    
    sheet['A6'] = "High Temp (°F):"
    sheet['B6'] = ""
    
    sheet['C6'] = "Low Temp (°F):"
    sheet['D6'] = ""
    
    sheet['E6'] = "Precipitation (in):"
    sheet['F6'] = ""
    
    sheet['G6'] = "Conditions:"
    sheet['H6'] = ""
    
    # Weather conditions dropdown
    weather_conditions = DataValidation(type="list", formula1='"Clear,Partly Cloudy,Cloudy,Rain,Snow,Fog,Windy,Stormy"')
    sheet.add_data_validation(weather_conditions)
    weather_conditions.add('H6')
    
    sheet['A7'] = "Weather Impact:"
    sheet['B7'] = "No"
    
    # Weather impact dropdown
    weather_impact = DataValidation(type="list", formula1='"Yes,No"')
    sheet.add_data_validation(weather_impact)
    weather_impact.add('B7')
    
    sheet['C7'] = "Description:"
    sheet.merge_cells('D7:H7')
    sheet['D7'] = ""
    
    # Manpower section
    sheet.merge_cells('A9:H9')
    sheet['A9'] = "MANPOWER"
    sheet['A9'].font = subheader_font
    sheet['A9'].fill = subheader_fill
    sheet['A9'].alignment = Alignment(horizontal='center')
    
    # Manpower table headers
    manpower_headers = ["Trade", "Company", "Workers", "Hours", "Total Man-Hours", "Work Area", "Notes"]
    for col, header in enumerate(manpower_headers, start=1):
        cell = sheet.cell(row=10, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
    
    # Add rows for manpower entries
    for row in range(11, 16):
        for col in range(1, 8):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            # Add formula for Total Man-Hours calculation
            if col == 5:
                cell.value = f"=C{row}*D{row}"
    
    # Trade dropdown
    trades = DataValidation(type="list", formula1='"Carpenter,Electrician,Plumber,HVAC,Laborer,Operator,Mason,Painter,Roofer,Other"')
    sheet.add_data_validation(trades)
    for row in range(11, 16):
        trades.add(f'A{row}')
    
    # Equipment section
    sheet.merge_cells('A17:H17')
    sheet['A17'] = "EQUIPMENT ON SITE"
    sheet['A17'].font = subheader_font
    sheet['A17'].fill = subheader_fill
    sheet['A17'].alignment = Alignment(horizontal='center')
    
    # Equipment table headers
    equipment_headers = ["Equipment Type", "ID/Number", "Quantity", "Hours Used", "Status", "Owner", "Notes"]
    for col, header in enumerate(equipment_headers, start=1):
        cell = sheet.cell(row=18, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
    
    # Add rows for equipment entries
    for row in range(19, 24):
        for col in range(1, 8):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
    
    # Equipment status dropdown
    equipment_status = DataValidation(type="list", formula1='"Active,Idle,Maintenance,Removed"')
    sheet.add_data_validation(equipment_status)
    for row in range(19, 24):
        equipment_status.add(f'E{row}')
    
    # Material Deliveries section
    sheet.merge_cells('A25:H25')
    sheet['A25'] = "MATERIAL DELIVERIES"
    sheet['A25'].font = subheader_font
    sheet['A25'].fill = subheader_fill
    sheet['A25'].alignment = Alignment(horizontal='center')
    
    # Material deliveries table headers
    materials_headers = ["Material", "Supplier", "Quantity", "Units", "Delivery Ticket #", "QC Check", "Storage Location", "Notes"]
    for col, header in enumerate(materials_headers, start=1):
        cell = sheet.cell(row=26, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
    
    # Add rows for material entries
    for row in range(27, 32):
        for col in range(1, 9):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
    
    # QC Check dropdown
    qc_check = DataValidation(type="list", formula1='"Yes,No,N/A"')
    sheet.add_data_validation(qc_check)
    for row in range(27, 32):
        qc_check.add(f'F{row}')
    
    # Visitors section
    sheet.merge_cells('A33:H33')
    sheet['A33'] = "VISITORS"
    sheet['A33'].font = subheader_font
    sheet['A33'].fill = subheader_fill
    sheet['A33'].alignment = Alignment(horizontal='center')
    
    # Visitors table headers
    visitors_headers = ["Name", "Company", "Purpose", "Time In", "Time Out", "Notes"]
    for col, header in enumerate(visitors_headers, start=1):
        cell = sheet.cell(row=34, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
    
    # Add rows for visitor entries
    for row in range(35, 39):
        for col in range(1, 7):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
    
    # Progress section
    sheet.merge_cells('A40:H40')
    sheet['A40'] = "WORK COMPLETED"
    sheet['A40'].font = subheader_font
    sheet['A40'].fill = subheader_fill
    sheet['A40'].alignment = Alignment(horizontal='center')
    
    sheet.merge_cells('A41:H43')
    sheet['A41'].border = border
    
    # Delays and Issues section
    sheet.merge_cells('A45:H45')
    sheet['A45'] = "DELAYS AND ISSUES"
    sheet['A45'].font = subheader_font
    sheet['A45'].fill = subheader_fill
    sheet['A45'].alignment = Alignment(horizontal='center')
    
    sheet['A46'] = "Delays Encountered:"
    sheet['B46'] = "No"
    
    # Delays dropdown
    delays = DataValidation(type="list", formula1='"Yes,No"')
    sheet.add_data_validation(delays)
    delays.add('B46')
    
    sheet['A47'] = "Description:"
    sheet.merge_cells('B47:H49')
    sheet['B47'].border = border
    
    # Photos section
    sheet.merge_cells('A51:H51')
    sheet['A51'] = "PROGRESS PHOTOS"
    sheet['A51'].font = subheader_font
    sheet['A51'].fill = subheader_fill
    sheet['A51'].alignment = Alignment(horizontal='center')
    
    # Photos table headers
    photos_headers = ["Photo Ref #", "Description", "Location", "Type"]
    for col, header in enumerate(photos_headers, start=1):
        cell = sheet.cell(row=52, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
    
    # Add rows for photo entries
    for row in range(53, 57):
        for col in range(1, 5):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
        # Formula for auto-generating photo reference numbers
        sheet.cell(row=row, column=1).value = f"=IF(B{row}=\"\",\"\",CONCATENATE(\"P-\",TEXT(B3,\"yyyymmdd\"),\"-\",{row-52}))"
    
    # Photo type dropdown
    photo_type = DataValidation(type="list", formula1='"Progress,QC,Safety,Weather,Issue"')
    sheet.add_data_validation(photo_type)
    for row in range(53, 57):
        photo_type.add(f'D{row}')
    
    # Signature section
    sheet.merge_cells('A59:H59')
    sheet['A59'] = "SIGN OFF"
    sheet['A59'].font = subheader_font
    sheet['A59'].fill = subheader_fill
    sheet['A59'].alignment = Alignment(horizontal='center')
    
    sheet['A60'] = "Prepared By:"
    sheet.merge_cells('B60:C60')
    sheet['B60'] = ""
    sheet['B60'].border = Border(bottom=Side(style='thin'))
    
    sheet['D60'] = "Title:"
    sheet.merge_cells('E60:F60')
    sheet['E60'] = ""
    sheet['E60'].border = Border(bottom=Side(style='thin'))
    
    sheet['G60'] = "Date:"
    sheet['H60'] = "=B3"
    sheet['H60'].number_format = "mm/dd/yyyy"
    
    # Submit button note
    sheet.merge_cells('A62:H62')
    sheet['A62'] = "NOTE: Click the 'Submit Report' button to save this report to the Reports Log."
    sheet['A62'].font = Font(italic=True)
    sheet['A62'].alignment = Alignment(horizontal='center')

def configure_reports_log(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:J1')
    sheet['A1'] = "DAILY REPORTS LOG"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Date", "Report #", "Superintendent", "Weather", "Manpower Count", "Key Activities", "Delays", "Safety Incidents", "Photos", "Link"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set up conditional formatting for the log
    # In a real implementation, we would add conditional formatting rules here
    
    # Add a few sample rows with borders
    for row in range(3, 100):
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            cell.border = border

def configure_weather_data(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "WEATHER DATA TRACKING"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Date", "High Temp (°F)", "Low Temp (°F)", "Precipitation (in)", "Weather Impact", "Conditions", "Description"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add rows with borders
    for row in range(3, 100):
        for col in range(1, 8):
            cell = sheet.cell(row=row, column=col)
            cell.border = border

def configure_manpower_tracking(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:H1')
    sheet['A1'] = "MANPOWER TRACKING"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Date", "Trade", "Company", "Workers", "Hours", "Total Man-Hours", "Work Area", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add formula for total man-hours
    sheet['F3'] = "=D3*E3"
    
    # Add rows with borders and formulas
    for row in range(3, 100):
        for col in range(1, 9):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            if col == 6:
                cell.value = f"=D{row}*E{row}"

def configure_equipment_log(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "EQUIPMENT LOG"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Equipment Type", "ID/Number", "Owner", "Date Arrived", "Date Removed", "Days On Site", "Status", "Hours Used", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Formula for days on site calculation
    sheet['F3'] = "=IF(AND(D3<>\"\",E3<>\"\"),E3-D3,IF(D3<>\"\",TODAY()-D3,\"\"))"
    
    # Add rows with borders and formulas
    for row in range(3, 100):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            if col == 6:
                cell.value = f"=IF(AND(D{row}<>\"\",E{row}<>\"\"),E{row}-D{row},IF(D{row}<>\"\",TODAY()-D{row},\"\"))"

def configure_material_deliveries(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "MATERIAL DELIVERIES"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Date", "Material", "Supplier", "Quantity", "Units", "PO Number", "Delivery Ticket #", "QC Check", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add rows with borders
    for row in range(3, 100):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            cell.border = border

def configure_photos_index(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:H1')
    sheet['A1'] = "PROGRESS PHOTOS INDEX"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Photo Ref #", "Date", "Description", "Location", "Direction", "Category", "File Name", "Link"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Formula for photo reference number
    sheet['A3'] = "=IF(B3=\"\",\"\",CONCATENATE(\"P-\",TEXT(B3,\"yyyymmdd\"),\"-1\"))"
    
    # Add rows with borders and formulas
    for row in range(3, 100):
        for col in range(1, 9):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            if col == 1 and row > 3:
                cell.value = f"=IF(B{row}=\"\",\"\",CONCATENATE(\"P-\",TEXT(B{row},\"yyyymmdd\"),\"-{row-2}\"))"

def configure_schedule_tracking(sheet, header_font, header_fill, subheader_font, subheader_fill, border):
    # Title
    sheet.merge_cells('A1:K1')
    sheet['A1'] = "SCHEDULE DELAYS TRACKING"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Table headers
    headers = ["Date Reported", "Activity", "Original Start", "Original Finish", "Revised Start", "Revised Finish", "Delay (Days)", 
               "Cause", "Critical Path", "Responsibility", "Mitigation"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Formula for delay calculation
    sheet['G3'] = "=IF(AND(C3<>\"\",D3<>\"\",E3<>\"\",F3<>\"\"),(F3-E3)-(D3-C3),\"\")"
    
    # Add rows with borders and formulas
    for row in range(3, 100):
        for col in range(1, 12):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            if col == 7:
                cell.value = f"=IF(AND(C{row}<>\"\",D{row}<>\"\",E{row}<>\"\",F{row}<>\"\"),(F{row}-E{row})-(D{row}-C{row}),\"\")"

if __name__ == "__main__":
    create_construction_daily_report()
