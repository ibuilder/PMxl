import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_construction_safety_program():
    """
    Creates an Excel workbook for a construction site safety program with
    multiple sheets for different safety forms and logs.
    """
    # Create Excel writer
    excel_file = 'Construction_Safety_Program.xlsx'
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    
    # Create each worksheet
    create_cover_sheet(writer)
    create_employee_orientation(writer)
    create_job_hazard_analysis(writer)
    create_pretask_plan(writer)
    create_inspection_log(writer)
    create_observation_report(writer)
    
    # Save the Excel file
    writer.close()
    
    print(f"Excel file '{excel_file}' has been created successfully!")
    return excel_file


def create_cover_sheet(writer):
    """Creates the cover sheet with instructions"""
    # Create cover sheet data
    cover_data = [
        ["CONSTRUCTION SITE SAFETY PROGRAM", "", "", ""],
        ["", "", "", ""],
        ["Company Name:", "", "", ""],
        ["Project Name:", "", "", ""],
        ["Project Location:", "", "", ""],
        ["Project Manager:", "", "", ""],
        ["Safety Manager:", "", "", ""],
        ["Emergency Contact:", "", "", ""],
        ["", "", "", ""],
        ["WORKBOOK CONTENTS:", "", "", ""],
        ["", "", "", ""],
        ["1. Employee Orientation Form", "Document used to record safety orientation for new employees", "", ""],
        ["2. Job Hazard Analysis (JHA)", "Used to identify and control hazards for specific job tasks", "", ""],
        ["3. Pre-Task Plan", "Daily planning tool to identify hazards before starting work", "", ""],
        ["4. Inspection Log", "Record of safety inspections conducted on the jobsite", "", ""],
        ["5. Observation Report", "Form to document safety observations and feedback", "", ""],
        ["", "", "", ""],
        ["INSTRUCTIONS:", "", "", ""],
        ["", "", "", ""],
        ["1. Complete all applicable forms as required by your safety program.", "", "", ""],
        ["2. Maintain records of all completed forms for the duration of the project.", "", "", ""],
        ["3. Review safety documentation regularly to identify trends and areas for improvement.", "", "", ""],
        ["4. Share relevant safety information with all project personnel.", "", "", ""],
        ["5. Update forms as needed to address changing conditions or requirements.", "", "", ""],
        ["", "", "", ""],
        ["REMEMBER: SAFETY IS EVERYONE'S RESPONSIBILITY", "", "", ""]
    ]
    
    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(cover_data)
    df.to_excel(writer, sheet_name='Cover Sheet', header=False, index=False)
    
    # Get the worksheet to apply formatting
    worksheet = writer.sheets['Cover Sheet']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    
    # Format the header
    worksheet['A1'].font = Font(bold=True, size=16)
    
    # Format section titles
    for cell in [worksheet['A10'], worksheet['A18'], worksheet['A27']]:
        cell.font = Font(bold=True, size=12)
    
    # Apply formatting to the whole sheet
    format_worksheet(worksheet, len(cover_data), 4)


def create_employee_orientation(writer):
    """Creates the employee orientation form"""
    # Create orientation form data
    orientation_data = [
        ["CONSTRUCTION SITE SAFETY PROGRAM", "", "", "", ""],
        ["EMPLOYEE ORIENTATION FORM", "", "", "", ""],
        ["", "", "", "", ""],
        ["Project Name:", "", "Project #:", "", ""],
        ["Project Location:", "", "", "", ""],
        ["Employee Name:", "", "Employee ID:", "", ""],
        ["Job Title:", "", "Start Date:", "", ""],
        ["Supervisor:", "", "", "", ""],
        ["", "", "", "", ""],
        ["ORIENTATION CHECKLIST", "YES", "NO", "N/A", "COMMENTS"],
        ["Company Safety Policy Reviewed", "", "", "", ""],
        ["Safety Rules and Procedures Explained", "", "", "", ""],
        ["Personal Protective Equipment Requirements", "", "", "", ""],
        ["Emergency Procedures and Exit Routes", "", "", "", ""],
        ["First Aid Locations and Procedures", "", "", "", ""],
        ["Incident Reporting Procedures", "", "", "", ""],
        ["Hazard Communication Program", "", "", "", ""],
        ["Confined Space Procedures", "", "", "", ""],
        ["Fall Protection Requirements", "", "", "", ""],
        ["Lockout/Tagout Procedures", "", "", "", ""],
        ["Fire Prevention and Protection", "", "", "", ""],
        ["Electrical Safety", "", "", "", ""],
        ["Hand and Power Tool Safety", "", "", "", ""],
        ["Heavy Equipment Safety", "", "", "", ""],
        ["Scaffolding and Ladder Safety", "", "", "", ""],
        ["", "", "", "", ""],
        ["Employee Comments/Questions:", "", "", "", ""],
        ["", "", "", "", ""],
        ["", "", "", "", ""],
        ["I acknowledge that I have received orientation on the above items and understand the safety requirements for this project.", "", "", "", ""],
        ["", "", "", "", ""],
        ["Employee Signature:", "", "Date:", "", ""],
        ["Orientation Conducted By:", "", "Date:", "", ""],
        ["Supervisor Signature:", "", "Date:", "", ""]
    ]
    
    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(orientation_data)
    df.to_excel(writer, sheet_name='Employee Orientation', header=False, index=False)
    
    # Get the worksheet to apply formatting
    worksheet = writer.sheets['Employee Orientation']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 10
    worksheet.column_dimensions['E'].width = 30
    
    # Format headers
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A2'].font = Font(bold=True, size=14)
    worksheet['A10'].font = Font(bold=True)
    
    # Format the checklist header row
    for col in range(1, 6):
        cell = worksheet.cell(row=10, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Apply formatting to the whole sheet
    format_worksheet(worksheet, len(orientation_data), 5)


def create_job_hazard_analysis(writer):
    """Creates the job hazard analysis form"""
    # Create JHA form data
    jha_data = [
        ["JOB HAZARD ANALYSIS (JHA) FORM", "", "", "", ""],
        ["", "", "", "", ""],
        ["Project Name:", "", "Project #:", "", ""],
        ["Location:", "", "Date:", "", ""],
        ["Task Description:", "", "", "", ""],
        ["Prepared By:", "", "Reviewed By:", "", ""],
        ["Required PPE:", "", "", "", ""],
        ["Required Tools/Equipment:", "", "", "", ""],
        ["Required Training:", "", "", "", ""],
        ["", "", "", "", ""],
        ["TASK STEPS", "POTENTIAL HAZARDS", "RISK LEVEL (H/M/L)", "CONTROL MEASURES", "RESPONSIBLE PERSON"],
        ["1.", "", "", "", ""],
        ["", "", "", "", ""],
        ["2.", "", "", "", ""],
        ["", "", "", "", ""],
        ["3.", "", "", "", ""],
        ["", "", "", "", ""],
        ["4.", "", "", "", ""],
        ["", "", "", "", ""],
        ["5.", "", "", "", ""],
        ["", "", "", "", ""],
        ["6.", "", "", "", ""],
        ["", "", "", "", ""],
        ["7.", "", "", "", ""],
        ["", "", "", "", ""],
        ["8.", "", "", "", ""],
        ["", "", "", "", ""],
        ["", "", "", "", ""],
        ["Emergency Procedures:", "", "", "", ""],
        ["", "", "", "", ""],
        ["APPROVALS", "", "", "", ""],
        ["Supervisor:", "", "Date:", "", ""],
        ["Safety Representative:", "", "Date:", "", ""],
        ["Project Manager:", "", "Date:", "", ""]
    ]
    
    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(jha_data)
    df.to_excel(writer, sheet_name='Job Hazard Analysis', header=False, index=False)
    
    # Get the worksheet to apply formatting
    worksheet = writer.sheets['Job Hazard Analysis']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 20
    
    # Format headers
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A11'].font = Font(bold=True)
    worksheet['A30'].font = Font(bold=True)
    
    # Format the task steps header row
    for col in range(1, 6):
        cell = worksheet.cell(row=11, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Apply formatting to the whole sheet
    format_worksheet(worksheet, len(jha_data), 5)


def create_pretask_plan(writer):
    """Creates the pre-task plan form"""
    # Create pre-task plan form data
    pretask_data = [
        ["DAILY PRE-TASK PLAN", "", "", "", ""],
        ["", "", "", "", ""],
        ["Project Name:", "", "Project #:", "", ""],
        ["Task Location:", "", "Date:", "", ""],
        ["Task Description:", "", "", "", ""],
        ["Supervisor:", "", "", "", ""],
        ["", "", "", "", ""],
        ["CREW MEMBERS (Print Name & Sign)", "", "", "", ""],
        ["1.", "", "5.", "", ""],
        ["2.", "", "6.", "", ""],
        ["3.", "", "7.", "", ""],
        ["4.", "", "8.", "", ""],
        ["", "", "", "", ""],
        ["HAZARD IDENTIFICATION & CONTROL MEASURES", "", "", "", ""],
        ["POTENTIAL HAZARDS", "YES/NO", "CONTROL MEASURES", "", ""],
        ["Falls > 6 feet", "", "", "", ""],
        ["Electrical Hazards", "", "", "", ""],
        ["Confined Spaces", "", "", "", ""],
        ["Hot Work", "", "", "", ""],
        ["Hazardous Materials", "", "", "", ""],
        ["Heavy Equipment", "", "", "", ""],
        ["Overhead Work", "", "", "", ""],
        ["Excavation/Trenching", "", "", "", ""],
        ["Noise > 85 dBA", "", "", "", ""],
        ["Manual Lifting", "", "", "", ""],
        ["Weather Conditions", "", "", "", ""],
        ["Other:", "", "", "", ""],
        ["", "", "", "", ""],
        ["REQUIRED PPE", "YES/NO", "TOOLS & EQUIPMENT NEEDED", "YES/NO", ""],
        ["Hard Hat", "", "Ladders", "", ""],
        ["Safety Glasses", "", "Scaffolding", "", ""],
        ["Safety Vest", "", "Power Tools", "", ""],
        ["Safety Footwear", "", "Hand Tools", "", ""],
        ["Gloves", "", "Heavy Equipment", "", ""],
        ["Hearing Protection", "", "Lifts", "", ""],
        ["Respiratory Protection", "", "Fall Protection", "", ""],
        ["Face Shield", "", "Fire Extinguisher", "", ""],
        ["Other:", "", "Other:", "", ""],
        ["", "", "", "", ""],
        ["PERMITS REQUIRED", "YES/NO", "", "", ""],
        ["Hot Work", "", "", "", ""],
        ["Confined Space", "", "", "", ""],
        ["Excavation", "", "", "", ""],
        ["Lockout/Tagout", "", "", "", ""],
        ["Other:", "", "", "", ""],
        ["", "", "", "", ""],
        ["EMERGENCY RESPONSE PLAN", "", "", "", ""],
        ["Assembly Point:", "", "", "", ""],
        ["First Aid Location:", "", "", "", ""],
        ["Emergency Contact:", "", "", "", ""],
        ["", "", "", "", ""],
        ["Supervisor Signature:", "", "Date:", "", ""]
    ]
    
    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(pretask_data)
    df.to_excel(writer, sheet_name='Pre-Task Plan', header=False, index=False)
    
    # Get the worksheet to apply formatting
    worksheet = writer.sheets['Pre-Task Plan']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 25
    
    # Format headers
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A8'].font = Font(bold=True)
    worksheet['A14'].font = Font(bold=True)
    worksheet['A15'].font = Font(bold=True)
    worksheet['A29'].font = Font(bold=True)
    worksheet['A41'].font = Font(bold=True)
    worksheet['A49'].font = Font(bold=True)
    
    # Format header rows
    for row_num in [15, 29, 41]:
        for col in range(1, 6):
            if col <= 3 or (row_num == 29 and col <= 4):
                cell = worksheet.cell(row=row_num, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Apply formatting to the whole sheet
    format_worksheet(worksheet, len(pretask_data), 5)


def create_inspection_log(writer):
    """Creates the inspection log"""
    # Create inspection log data
    inspection_data = [
        ["SAFETY INSPECTION LOG", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Project Name:", "", "Project #:", "", "", "", ""],
        ["Project Location:", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["DATE", "AREA INSPECTED", "INSPECTED BY", "FINDINGS/HAZARDS", "CORRECTIVE ACTIONS", "RESPONSIBLE PERSON", "COMPLETION DATE"],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Project Manager Review:", "", "Date:", "", "", "", ""],
        ["Safety Manager Review:", "", "Date:", "", "", "", ""]
    ]
    
    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(inspection_data)
    df.to_excel(writer, sheet_name='Inspection Log', header=False, index=False)
    
    # Get the worksheet to apply formatting
    worksheet = writer.sheets['Inspection Log']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 30
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 15
    
    # Format headers
    worksheet['A1'].font = Font(bold=True, size=14)
    
    # Format the header row
    for col in range(1, 8):
        cell = worksheet.cell(row=6, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Apply formatting to the whole sheet
    format_worksheet(worksheet, len(inspection_data), 7)


def create_observation_report(writer):
    """Creates the observation report form"""
    # Create observation report data
    observation_data = [
        ["SAFETY OBSERVATION REPORT", "", "", "", ""],
        ["", "", "", "", ""],
        ["Project Name:", "", "Project #:", "", ""],
        ["Project Location:", "", "Date:", "", ""],
        ["Observer Name:", "", "Position:", "", ""],
        ["", "", "", "", ""],
        ["Type of Observation:", "", "", "", ""],
        ["☐ Planned   ☐ Unplanned", "", "", "", ""],
        ["Area/Location:", "", "", "", ""],
        ["Task Observed:", "", "", "", ""],
        ["", "", "", "", ""],
        ["SAFE PRACTICES OBSERVED", "", "", "", ""],
        ["", "", "", "", ""],
        ["1.", "", "", "", ""],
        ["2.", "", "", "", ""],
        ["3.", "", "", "", ""],
        ["4.", "", "", "", ""],
        ["", "", "", "", ""],
        ["AT-RISK BEHAVIORS/CONDITIONS OBSERVED", "CORRECTIVE ACTION TAKEN", "FOLLOW-UP REQUIRED", "RESPONSIBLE PERSON", "DUE DATE"],
        ["1.", "", "", "", ""],
        ["2.", "", "", "", ""],
        ["3.", "", "", "", ""],
        ["4.", "", "", "", ""],
        ["", "", "", "", ""],
        ["OBSERVATION CATEGORIES", "YES", "NO", "N/A", "COMMENTS"],
        ["1. PPE Used Properly", "", "", "", ""],
        ["2. Tools & Equipment in Good Condition", "", "", "", ""],
        ["3. Proper Body Positioning/Mechanics", "", "", "", ""],
        ["4. Work Area Clean & Orderly", "", "", "", ""],
        ["5. Procedures Being Followed", "", "", "", ""],
        ["6. Hazard Controls in Place", "", "", "", ""],
        ["7. Communication Between Workers", "", "", "", ""],
        ["8. Environmental Conditions Addressed", "", "", "", ""],
        ["", "", "", "", ""],
        ["Feedback Provided To:", "", "", "", ""],
        ["Feedback Summary:", "", "", "", ""],
        ["", "", "", "", ""],
        ["Observer Signature:", "", "Date:", "", ""],
        ["Supervisor Review:", "", "Date:", "", ""],
        ["Safety Manager Review:", "", "Date:", "", ""]
    ]
    
    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(observation_data)
    df.to_excel(writer, sheet_name='Observation Report', header=False, index=False)
    
    # Get the worksheet to apply formatting
    worksheet = writer.sheets['Observation Report']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 15
    
    # Format headers
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A12'].font = Font(bold=True)
    worksheet['A19'].font = Font(bold=True)
    worksheet['A26'].font = Font(bold=True)
    
    # Format header rows
    for row_num in [19, 26]:
        for col in range(1, 6):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Apply formatting to the whole sheet
    format_worksheet(worksheet, len(observation_data), 5)


def format_worksheet(worksheet, num_rows, num_cols):
    """Apply general formatting to a worksheet"""
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders and alignment to all cells with data
    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Make field labels bold
            if col == 1 and ":" in str(cell.value):
                cell.font = Font(bold=True)


if __name__ == "__main__":
    create_construction_safety_program()
