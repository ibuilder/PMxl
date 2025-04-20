import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def create_meeting_workbook(filename="Meeting_Management_Workbook.xlsx"):
    """
    Create an Excel workbook with three sheets for meeting management:
    1. Meeting Log
    2. Meeting Agenda
    3. Meeting Minutes
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create Meeting Log sheet
    meeting_log = wb.create_sheet("Meeting Log")
    
    # Title row
    meeting_log.merge_cells('A1:F1')
    meeting_log['A1'] = 'Meeting Log'
    meeting_log['A1'].font = Font(bold=True, size=14)
    meeting_log['A1'].alignment = Alignment(horizontal='center')
    
    # Header row
    headers = ['Meeting ID', 'Date', 'Time', 'Location', 'Attendees', 'Purpose']
    for col_num, header in enumerate(headers, 1):
        cell = meeting_log.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Sample data
    sample_data = [
        ['ML001', '2025-04-20', '09:00-10:00', 'Conference Room A', 'Team Members', 'Project Kickoff'],
        ['ML002', '2025-04-27', '14:00-15:00', 'Virtual Meeting', 'Department Heads', 'Monthly Review']
    ]
    
    for row_num, row_data in enumerate(sample_data, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            meeting_log.cell(row=row_num, column=col_num).value = cell_value
    
    # Set column widths
    col_widths = [15, 15, 20, 20, 25, 25]
    for i, width in enumerate(col_widths, 1):
        meeting_log.column_dimensions[get_column_letter(i)].width = width
    
    # Create Meeting Agenda sheet
    meeting_agenda = wb.create_sheet("Meeting Agenda")
    
    # Title row
    meeting_agenda.merge_cells('A1:E1')
    meeting_agenda['A1'] = 'Meeting Agenda'
    meeting_agenda['A1'].font = Font(bold=True, size=14)
    meeting_agenda['A1'].alignment = Alignment(horizontal='center')
    
    # Meeting details
    agenda_fields = [
        ['Meeting Title:', '', '', '', ''],
        ['Date:', '', '', '', ''],
        ['Time:', '', '', '', ''],
        ['Location:', '', '', '', ''],
        ['Meeting Called By:', '', '', '', ''],
        ['Attendees:', '', '', '', ''],
        ['', '', '', '', ''],
        ['Objective:', '', '', '', ''],
        ['', '', '', '', ''],
        ['Agenda Items:', 'Duration', 'Presenter', 'Notes', ''],
    ]
    
    for row_num, row_data in enumerate(agenda_fields, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            meeting_agenda.cell(row=row_num, column=col_num).value = cell_value
            if row_num == 12:  # Header row for agenda items
                meeting_agenda.cell(row=row_num, column=col_num).font = Font(bold=True)
    
    # Merge cells for fields
    merge_ranges = ['B3:E3', 'B4:E4', 'B5:E5', 'B6:E6', 'B7:E7', 'B8:E8', 'B10:E10']
    for cell_range in merge_ranges:
        meeting_agenda.merge_cells(cell_range)
    
    # Agenda items
    for i in range(1, 6):
        meeting_agenda.cell(row=12+i, column=1).value = f"{i}."
    
    # Additional sections
    meeting_agenda.cell(row=19, column=1).value = 'Pre-meeting Preparation:'
    meeting_agenda.merge_cells('B19:E19')
    
    meeting_agenda.cell(row=21, column=1).value = 'Additional Information:'
    meeting_agenda.merge_cells('B21:E21')
    
    # Set column widths
    col_widths = [20, 30, 15, 15, 30]
    for i, width in enumerate(col_widths, 1):
        meeting_agenda.column_dimensions[get_column_letter(i)].width = width
    
    # Create Meeting Minutes sheet
    meeting_minutes = wb.create_sheet("Meeting Minutes")
    
    # Title row
    meeting_minutes.merge_cells('A1:E1')
    meeting_minutes['A1'] = 'Meeting Minutes'
    meeting_minutes['A1'].font = Font(bold=True, size=14)
    meeting_minutes['A1'].alignment = Alignment(horizontal='center')
    
    # Meeting details
    minutes_fields = [
        ['Meeting Title:', '', '', '', ''],
        ['Date:', '', '', '', ''],
        ['Time:', '', '', '', ''],
        ['Location:', '', '', '', ''],
        ['Attendees Present:', '', '', '', ''],
        ['Attendees Absent:', '', '', '', ''],
        ['', '', '', '', ''],
        ['Agenda Items', 'Discussion Points', 'Action Items', 'Person Responsible', 'Deadline']
    ]
    
    for row_num, row_data in enumerate(minutes_fields, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            meeting_minutes.cell(row=row_num, column=col_num).value = cell_value
            if row_num == 10:  # Header row for minutes items
                meeting_minutes.cell(row=row_num, column=col_num).font = Font(bold=True)
    
    # Merge cells for fields
    merge_ranges = ['B3:E3', 'B4:E4', 'B5:E5', 'B6:E6', 'B7:E7', 'B8:E8']
    for cell_range in merge_ranges:
        meeting_minutes.merge_cells(cell_range)
    
    # Agenda items and discussion
    current_row = 11
    for i in range(1, 6):
        meeting_minutes.cell(row=current_row, column=1).value = f"{i}."
        current_row += 2
    
    # Next meeting section
    next_meeting_row = current_row + 1
    meeting_minutes.cell(row=next_meeting_row, column=1).value = 'Next Meeting:'
    meeting_minutes.cell(row=next_meeting_row+1, column=1).value = 'Date:'
    meeting_minutes.cell(row=next_meeting_row+2, column=1).value = 'Time:'
    meeting_minutes.cell(row=next_meeting_row+3, column=1).value = 'Location:'
    meeting_minutes.cell(row=next_meeting_row+4, column=1).value = 'Agenda Items:'
    
    # Set column widths
    col_widths = [20, 30, 20, 20, 15]
    for i, width in enumerate(col_widths, 1):
        meeting_minutes.column_dimensions[get_column_letter(i)].width = width
    
    # Save the workbook
    wb.save(filename)
    print(f"Meeting workbook created successfully: {filename}")

if __name__ == "__main__":
    create_meeting_workbook()
