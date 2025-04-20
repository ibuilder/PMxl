import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_construction_submittal_workbook(filename="Construction_Submittal_Workbook.xlsx"):
    """
    Creates an Excel workbook for construction submittal tracking with multiple sheets:
    - Submittal Log
    - Package Template
    - Item Form
    - Instructions
    
    Args:
        filename (str): The name of the Excel file to create
    """
    # Create a workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create all sheets
    create_submittal_log(wb)
    create_package_template(wb)
    create_item_form(wb)
    create_instructions(wb)
    
    # Save the workbook
    wb.save(filename)
    print(f"Workbook saved as {filename}")

def apply_header_style(cell):
    """Apply styling for headers"""
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    thin_border = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

def apply_subheader_style(cell):
    """Apply styling for subheaders"""
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def create_submittal_log(wb):
    """Create the Submittal Log sheet"""
    ws = wb.create_sheet(title="Submittal Log")
    
    # Set column widths
    column_widths = [12, 15, 25, 15, 20, 15, 15, 15, 15, 25]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Title
    ws['A1'] = "CONSTRUCTION SUBMITTAL LOG"
    ws.merge_cells('A1:J1')
    apply_header_style(ws['A1'])
    
    # Project info
    ws['A2'] = "Project Name:"
    ws['B2'] = "ENTER PROJECT NAME"
    ws.merge_cells('B2:E2')
    ws['G2'] = "Project No.:"
    ws['H2'] = "ENTER PROJECT NO."
    ws.merge_cells('H2:J2')
    
    # Contractor info
    ws['A3'] = "Contractor:"
    ws['B3'] = "ENTER CONTRACTOR NAME"
    ws.merge_cells('B3:E3')
    ws['G3'] = "Updated:"
    ws['H3'] = datetime.now().strftime("%m/%d/%Y")
    ws.merge_cells('H3:J3')
    
    # Column headers
    headers = ["Submittal No.", "Spec Section", "Description", "Submittal Type", 
               "Subcontractor", "Required Date", "Date Received", "Status", "Returned Date", "Remarks"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_header_style(cell)
    
    # Sample data
    sample_data = [
        ["001", "03 30 00", "Concrete Mix Design", "Product Data", "ABC Concrete", "05/01/2025", "04/15/2025", "Approved", "04/20/2025", ""],
        ["002", "04 20 00", "Masonry Units", "Samples", "XYZ Masonry", "05/15/2025", "04/22/2025", "Approved as Noted", "04/27/2025", "Resubmit color samples"],
        ["003", "05 12 00", "Structural Steel", "Shop Drawings", "Steel Fabricators Inc.", "05/30/2025", "", "Pending", "", ""],
        ["004", "07 21 00", "Insulation", "Product Data", "Insulation Co.", "06/15/2025", "", "Pending", "", ""],
        ["005", "08 11 13", "Hollow Metal Doors", "Shop Drawings", "Door Suppliers LLC", "06/30/2025", "", "Pending", "", ""]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 6):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add data validation for Status column
    status_validation = DataValidation(type="list", formula1='"Pending,In Review,Approved,Approved as Noted,Revise and Resubmit,Rejected,For Information Only"')
    ws.add_data_validation(status_validation)
    status_validation.add('H7:H100')
    
    # Add empty rows for additional entries
    for row in range(11, 51):
        ws.cell(row=row, column=1)

def create_package_template(wb):
    """Create the Submittal Package Template sheet"""
    ws = wb.create_sheet(title="Package Template")
    
    # Set column widths
    column_widths = [15, 25, 20, 15, 20, 20]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Title
    ws['A1'] = "SUBMITTAL PACKAGE COVER SHEET"
    ws.merge_cells('A1:F1')
    apply_header_style(ws['A1'])
    
    # Project info
    ws['A3'] = "Project Name:"
    ws['B3'] = "ENTER PROJECT NAME"
    ws.merge_cells('B3:C3')
    ws['D3'] = "Project No.:"
    ws['E3'] = "ENTER PROJECT NO."
    ws.merge_cells('E3:F3')
    
    # Contractor info
    ws['A4'] = "Contractor:"
    ws['B4'] = "ENTER CONTRACTOR NAME"
    ws.merge_cells('B4:C4')
    ws['D4'] = "Date:"
    ws['E4'] = datetime.now().strftime("%m/%d/%Y")
    ws.merge_cells('E4:F4')
    
    # Package info section
    ws['A6'] = "SUBMITTAL PACKAGE INFORMATION"
    ws.merge_cells('A6:F6')
    apply_subheader_style(ws['A6'])
    
    ws['A7'] = "Package No.:"
    ws['D7'] = "Specification Section(s):"
    
    ws['A8'] = "Package Description:"
    ws.merge_cells('B8:F8')
    
    ws['A9'] = "Submitted By:"
    ws['D9'] = "Title:"
    
    # Items section
    ws['A11'] = "SUBMITTAL ITEMS INCLUDED IN THIS PACKAGE"
    ws.merge_cells('A11:F11')
    apply_subheader_style(ws['A11'])
    
    # Column headers for items
    item_headers = ["Item No.", "Description", "Type", "No. of Copies", "Spec Reference", "Remarks"]
    for col, header in enumerate(item_headers, 1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_header_style(cell)
    
    # Empty rows for items
    for row in range(13, 18):
        ws.cell(row=row, column=1, value=row-12)
    
    # Review section
    ws['A19'] = "REVIEW ACTION"
    ws.merge_cells('A19:F19')
    apply_subheader_style(ws['A19'])
    
    # Checkboxes for review
    ws['A20'] = "□ APPROVED"
    ws['B20'] = "□ APPROVED AS NOTED"
    ws['C20'] = "□ REVISE AND RESUBMIT"
    ws['D20'] = "□ REJECTED"
    ws['E20'] = "□ FOR INFORMATION ONLY"
    
    # Comments section
    ws['A22'] = "Comments:"
    ws.merge_cells('B22:F22')
    ws.merge_cells('A23:F25')
    
    # Reviewer info
    ws['A26'] = "Reviewed By:"
    ws['D26'] = "Date:"

def create_item_form(wb):
    """Create the Submittal Item Form sheet"""
    ws = wb.create_sheet(title="Item Form")
    
    # Set column widths
    column_widths = [15, 20, 20, 15, 20]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Title
    ws['A1'] = "SUBMITTAL ITEM FORM"
    ws.merge_cells('A1:E1')
    apply_header_style(ws['A1'])
    
    # Project info
    ws['A3'] = "Project Name:"
    ws['B3'] = "ENTER PROJECT NAME"
    ws.merge_cells('B3:C3')
    ws['D3'] = "Project No.:"
    ws['E3'] = "ENTER PROJECT NO."
    
    # Contractor info
    ws['A4'] = "Contractor:"
    ws['B4'] = "ENTER CONTRACTOR NAME"
    ws.merge_cells('B4:C4')
    ws['D4'] = "Date:"
    ws['E4'] = datetime.now().strftime("%m/%d/%Y")
    
    # Submittal info section
    ws['A6'] = "SUBMITTAL INFORMATION"
    ws.merge_cells('A6:E6')
    apply_subheader_style(ws['A6'])
    
    ws['A7'] = "Submittal No.:"
    ws['D7'] = "Spec Section:"
    
    ws['A8'] = "Description:"
    ws.merge_cells('B8:E8')
    
    ws['A9'] = "Submittal Type:"
    
    # Checkboxes for submittal type
    ws['A10'] = "□ Product Data"
    ws['B10'] = "□ Shop Drawings"
    ws['C10'] = "□ Samples"
    ws['D10'] = "□ Quality Control"
    ws['E10'] = "□ Other:___________"
    
    ws['A11'] = "Subcontractor:"
    ws.merge_cells('B11:E11')
    
    ws['A12'] = "Manufacturer:"
    ws.merge_cells('B12:E12')
    
    ws['A13'] = "Supplier:"
    ws.merge_cells('B13:E13')
    
    # Contractor review section
    ws['A15'] = "CONTRACTOR REVIEW"
    ws.merge_cells('A15:E15')
    apply_subheader_style(ws['A15'])
    
    # Checkboxes for contractor review
    ws['A16'] = "□ Approved"
    ws['B16'] = "□ Approved as Noted"
    ws['C16'] = "□ Revise and Resubmit"
    ws['D16'] = "□ Rejected"
    
    # Comments section for contractor
    ws['A17'] = "Comments:"
    ws.merge_cells('B17:E17')
    ws.merge_cells('A18:E19')
    
    # Reviewer info for contractor
    ws['A20'] = "Reviewer:"
    ws['D20'] = "Date:"
    
    # A/E review section
    ws['A22'] = "ARCHITECT/ENGINEER REVIEW"
    ws.merge_cells('A22:E22')
    apply_subheader_style(ws['A22'])
    
    # Checkboxes for A/E review
    ws['A23'] = "□ Approved"
    ws['B23'] = "□ Approved as Noted"
    ws['C23'] = "□ Revise and Resubmit"
    ws['D23'] = "□ Rejected"
    ws['E23'] = "□ For Information Only"
    
    # Comments section for A/E
    ws['A24'] = "Comments:"
    ws.merge_cells('B24:E24')
    ws.merge_cells('A25:E26')
    
    # Reviewer info for A/E
    ws['A27'] = "Reviewer:"
    ws['D27'] = "Date:"

def create_instructions(wb):
    """Create the Instructions sheet"""
    ws = wb.create_sheet(title="Instructions")
    
    # Set column widths
    column_widths = [25, 30, 20, 20]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Title
    ws['A1'] = "CONSTRUCTION SUBMITTAL TRACKING WORKBOOK - INSTRUCTIONS"
    ws.merge_cells('A1:D1')
    apply_header_style(ws['A1'])
    
    # Content
    instructions = [
        ["Overview:", "", "", ""],
        ["This workbook contains templates for tracking construction submittals, creating submittal packages, and documenting individual submittal items.", "", "", ""],
        ["", "", "", ""],
        ["Sheets Included:", "", "", ""],
        ["1. Submittal Log - For tracking all project submittals", "", "", ""],
        ["2. Package Template - Template for creating submittal package cover sheets", "", "", ""],
        ["3. Item Form - Template for individual submittal item documentation", "", "", ""],
        ["", "", "", ""],
        ["How to Use:", "", "", ""],
        ["Submittal Log:", "", "", ""],
        ["- Enter project information at the top", "", "", ""],
        ["- Assign sequential submittal numbers", "", "", ""],
        ["- Enter specification section references", "", "", ""],
        ["- Track status and dates for each submittal", "", "", ""],
        ["- Update the log regularly as submittals are processed", "", "", ""],
        ["", "", "", ""],
        ["Package Template:", "", "", ""],
        ["- Complete this form when bundling multiple related submittals", "", "", ""],
        ["- Assign a package number", "", "", ""],
        ["- List all submittal items included in the package", "", "", ""],
        ["- Attach this as a cover sheet to the physical submittal package", "", "", ""],
        ["", "", "", ""],
        ["Item Form:", "", "", ""],
        ["- Complete this form for each individual submittal item", "", "", ""],
        ["- Document contractor review before submission", "", "", ""],
        ["- Record architect/engineer review after return", "", "", ""],
        ["- Include with the actual submittal materials", "", "", ""],
        ["", "", "", ""],
        ["Submittal Status Options:", "", "", ""],
        ["- Pending: Not yet received from subcontractor/supplier", "", "", ""],
        ["- In Review: Received but not yet reviewed", "", "", ""],
        ["- Approved: Reviewed and accepted without changes", "", "", ""],
        ["- Approved as Noted: Approved with minor changes noted", "", "", ""],
        ["- Revise and Resubmit: Requires correction and resubmission", "", "", ""],
        ["- Rejected: Not acceptable, must be resubmitted", "", "", ""],
        ["- For Information Only: Provided for reference, no approval required", "", "", ""],
        ["", "", "", ""],
        ["Tips:", "", "", ""],
        ["- Make a copy of the template sheets as needed", "", "", ""],
        ["- Maintain consistent submittal numbering", "", "", ""],
        ["- Use clear, descriptive titles for submittals", "", "", ""],
        ["- Track resubmittals with the original number plus a revision suffix (e.g., 001-R1)", "", "", ""],
        ["- Set up automatic conditional formatting to highlight late or overdue submittals", "", "", ""]
    ]
    
    # Write instructions
    for row_idx, row_data in enumerate(instructions, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if value.startswith("Submittal Log:") or value.startswith("Package Template:") or value.startswith("Item Form:") or value.startswith("Tips:") or value.startswith("Submittal Status Options:"):
                cell.font = Font(bold=True)

if __name__ == "__main__":
    create_construction_submittal_workbook()
